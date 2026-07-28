#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Разбор таблицы _base_kg: уникальные ИНН из «Компании группы (ЕРЗ)».

Независимый скрипт: читает config.json (блок base_kg_inn_extract),
парсит ячейки с компаниями вида «Название (ИНН 123); …»,
собирает таблицу уникальных ИНН и сохраняет отдельный Excel.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

КОРЕНЬ: Path = Path(__file__).resolve().parent.parent
КОНФИГ_ПО_УМОЛЧАНИЮ: Path = КОРЕНЬ / "config.json"

logger: logging.Logger = logging.getLogger("base_kg_inn")


def настроить_логирование(
    logs_dir: Path,
    log_file_prefix: str,
    log_to_file: bool,
) -> Path | None:
    """Консоль + опциональный файл лога в logs_dir."""
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter(
        "%(asctime)s - [%(levelname)s] - %(message)s [def: %(funcName)s]",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(fmt)
    logger.addHandler(console)

    if not log_to_file:
        return None

    logs_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H")
    log_path = logs_dir / f"{log_file_prefix}_{stamp}.log"
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    return log_path


def загрузить_конфиг(path: Path) -> dict[str, Any]:
    """Чтение JSON-конфига."""
    if not path.exists():
        raise FileNotFoundError(f"config-json не найден: {path}")
    with path.open(encoding="utf-8") as f:
        data: dict[str, Any] = json.load(f)
    return data


def разрешить_путь(value: str, base_dir: Path) -> Path:
    """Относительный путь — от каталога config.json."""
    p = Path(value).expanduser()
    if p.is_absolute():
        return p.resolve()
    return (base_dir / p).resolve()


def с_таймштампом(path: Path, pattern: str = "%Y%m%d_%H%M%S") -> Path:
    """Добавить таймштамп перед расширением файла."""
    ts = datetime.now().strftime(pattern)
    if path.suffix:
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")
    return path.with_name(f"{path.name}_{ts}")


def найти_ссылку_таблицы(path: Path, table_name: str) -> tuple[str, tuple[int, int, int, int]]:
    """Найти лист и границы смарт-таблицы Excel."""
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            if table_name in ws.tables:
                ref = ws.tables[table_name].ref
                min_col, min_row, max_col, max_row = range_boundaries(ref)
                return ws.title, (min_col, min_row, max_col, max_row)
    finally:
        wb.close()
    raise ValueError(f"Таблица '{table_name}' не найдена в {path}")


def прочитать_смарт_таблицу(path: Path, table_name: str) -> list[dict[str, Any]]:
    """Прочитать смарт-таблицу как список словарей (заголовок → значение)."""
    sheet_title, (min_col, min_row, max_col, max_row) = найти_ссылку_таблицы(path, table_name)
    logger.info("Читаю таблицу %s с листа «%s» (%s)", table_name, sheet_title, path)

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_title]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            rows.append(list(row))
    finally:
        wb.close()

    if not rows:
        return []

    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        out.append(rec)
    logger.debug("Прочитано строк данных: %s", len(out))
    return out


def _экранировать_для_класса(символ: str) -> str:
    """Экранирование символа для использования в regex character class / alternation."""
    return re.escape(символ)


def построить_шаблон_разделителей(separators: list[str]) -> re.Pattern[str]:
    """Регулярка для разбиения списка компаний по настроенным разделителям."""
    if not separators:
        raise ValueError("company_separators не должен быть пустым")
    # Длинные разделители раньше коротких (на случай многосимвольных).
    parts = sorted({s for s in separators if s}, key=len, reverse=True)
    if not parts:
        raise ValueError("company_separators не содержит непустых значений")
    alt = "|".join(_экранировать_для_класса(s) for s in parts)
    return re.compile(alt)


def разобрать_компании_ерз(
    text: Any,
    inn_keyword: str,
    separators: list[str],
) -> list[tuple[str, str]]:
    """Извлечь пары (наименование, ИНН) из ячейки «Компании группы (ЕРЗ)».

    Ожидаемый фрагмент: «АО Пример (ИНН 1234567890)».
    Несколько компаний разделяются символами из separators.
    """
    if text is None:
        return []
    raw = str(text).strip()
    if not raw:
        return []

    keyword = (inn_keyword or "ИНН").strip()
    if not keyword:
        raise ValueError("inn_keyword не должен быть пустым")

    split_re = построить_шаблон_разделителей(separators)
    # Паттерн фрагмента: название + (КЛЮЧЕВОЕ_СЛОВО цифры)
    fragment_re = re.compile(
        rf"^(?P<name>.+?)\s*\(\s*{re.escape(keyword)}\s+(?P<inn>\d+)\s*\)\s*$",
        flags=re.IGNORECASE | re.DOTALL,
    )

    result: list[tuple[str, str]] = []
    for part in split_re.split(raw):
        fragment = part.strip()
        if not fragment:
            continue
        match = fragment_re.match(fragment)
        if not match:
            logger.debug("Пропуск неразборчивого фрагмента: %r", fragment)
            continue
        name = match.group("name").strip()
        inn = match.group("inn").strip()
        if name and inn:
            result.append((name, inn))
    return result


def имя_до_запятой(region: str) -> str:
    """Часть «Наименование, регион» до первой запятой (без города/доп. данных)."""
    text = region.strip()
    if not text:
        return ""
    if "," in text:
        return text.split(",", 1)[0].strip()
    return text


OUTPUT_COLUMN_KEYS: list[str] = [
    "inn",
    "company_name",
    "region_name",
    "region_count",
    "region_base_name",
    "region_base_count",
]


def агрегировать_по_инн(
    rows: list[dict[str, Any]],
    region_column: str,
    companies_column: str,
    inn_keyword: str,
    separators: list[str],
    joiner: str,
) -> list[dict[str, Any]]:
    """Собрать уникальные ИНН с агрегацией названий, регионов и урезанных имён."""
    # inn -> OrderedDict имён / регионов (сохраняем порядок первого появления)
    names_by_inn: OrderedDict[str, OrderedDict[str, None]] = OrderedDict()
    regions_by_inn: OrderedDict[str, OrderedDict[str, None]] = OrderedDict()

    for row in rows:
        region_raw = row.get(region_column)
        region = "" if region_raw is None else str(region_raw).strip()
        companies_raw = row.get(companies_column)
        pairs = разобрать_компании_ерз(companies_raw, inn_keyword, separators)
        for company_name, inn in pairs:
            if inn not in names_by_inn:
                names_by_inn[inn] = OrderedDict()
                regions_by_inn[inn] = OrderedDict()
            names_by_inn[inn][company_name] = None
            if region:
                regions_by_inn[inn][region] = None

    out: list[dict[str, Any]] = []
    for inn, names in names_by_inn.items():
        regions = list(regions_by_inn[inn].keys())
        base_names: OrderedDict[str, None] = OrderedDict()
        for region in regions:
            base = имя_до_запятой(region)
            if base:
                base_names[base] = None
        out.append(
            {
                "inn": inn,
                "company_name": joiner.join(names.keys()),
                "region_name": joiner.join(regions),
                "region_count": len(regions),
                "region_base_name": joiner.join(base_names.keys()),
                "region_base_count": len(base_names),
            }
        )
    logger.info("Уникальных ИНН: %s (из %s строк источника)", len(out), len(rows))
    return out


def применить_форматирование(
    ws: Any,
    headers: list[str],
    format_cfg: dict[str, Any],
    column_keys: list[str],
    column_specs: dict[str, dict[str, Any]],
) -> None:
    """Заголовки, автофильтр, freeze, ширины и выравнивание ячеек."""
    header_bold = bool(format_cfg.get("header_bold", True))
    header_center = bool(format_cfg.get("header_center", True))
    header_wrap = bool(format_cfg.get("header_wrap", True))
    freeze_rows = max(0, int(format_cfg.get("freeze_rows", 1)))
    freeze_cols = max(0, int(format_cfg.get("freeze_cols", 1)))
    data_v_center = bool(format_cfg.get("data_vertical_center", True))
    data_h_default = str(format_cfg.get("data_horizontal", "left"))

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.alignment = Alignment(
                horizontal="center" if header_center else "general",
                vertical="center",
                wrap_text=header_wrap,
            )
            if header_bold:
                cell.font = Font(bold=True)

    # Данные: вертикальный центр; горизонталь/wrap — из спецификации колонки.
    if ws.max_row >= 2:
        for col_idx, key in enumerate(column_keys, start=1):
            spec = column_specs.get(key, {})
            wrap = bool(spec.get("wrap", key not in {"inn", "region_count", "region_base_count"}))
            align_h = str(spec.get("align", data_h_default))
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(
                    horizontal=align_h,
                    vertical="center" if data_v_center else "bottom",
                    wrap_text=wrap,
                )

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    if freeze_rows > 0 or freeze_cols > 0:
        ws.freeze_panes = ws.cell(row=freeze_rows + 1, column=freeze_cols + 1)

    default_widths: dict[str, float] = {
        "inn": 30.0,
        "company_name": 70.0,
        "region_name": 100.0,
        "region_count": 10.0,
        "region_base_name": 90.0,
        "region_base_count": 10.0,
    }
    for col_idx, key in enumerate(column_keys, start=1):
        width = float(
            column_specs.get(key, {}).get("width", default_widths.get(key, 30.0))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    _ = headers


def сохранить_результат(
    rows: list[dict[str, Any]],
    output_path: Path,
    sheet_name: str,
    output_columns: dict[str, str],
    format_cfg: dict[str, Any],
) -> None:
    """Записать агрегированную таблицу в отдельный Excel-файл."""
    column_keys = list(OUTPUT_COLUMN_KEYS)
    headers = [str(output_columns.get(k, k)) for k in column_keys]
    column_specs: dict[str, dict[str, Any]] = dict(format_cfg.get("columns") or {})

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "_BASE_KG_INN"

    ws.append(headers)
    for row in rows:
        ws.append([row.get(k, "") for k in column_keys])

    применить_форматирование(ws, headers, format_cfg, column_keys, column_specs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Сохранено: %s (строк: %s)", output_path, len(rows))


def разрешить_настройки(args: argparse.Namespace) -> dict[str, Any]:
    """Собрать рабочие настройки из config + CLI-переопределений."""
    config_path = Path(args.config_json).expanduser().resolve()
    cfg = загрузить_конфиг(config_path)
    block: dict[str, Any] = dict(cfg.get("base_kg_inn_extract") or {})

    def pick(cli_val: Any, key: str, default: Any = None) -> Any:
        if cli_val is not None:
            return cli_val
        if key in block and block[key] is not None:
            return block[key]
        return default

    settings: dict[str, Any] = {
        "_config_dir": str(config_path.parent),
        "input_xlsx": pick(args.input_xlsx, "input_xlsx"),
        "output_xlsx": pick(args.output_xlsx, "output_xlsx", "output/base_kg_inn.xlsx"),
        "output_add_timestamp": bool(
            pick(args.output_add_timestamp, "output_add_timestamp", True)
        ),
        "output_timestamp_format": str(
            pick(None, "output_timestamp_format", "%Y%m%d_%H%M%S")
        ),
        "output_sheet": str(pick(args.output_sheet, "output_sheet", "_BASE_KG_INN")),
        "base_table": str(pick(args.base_table, "base_table", "_base_kg")),
        "region_column": str(
            pick(args.region_column, "region_column", "Наименование, регион")
        ),
        "companies_column": str(
            pick(args.companies_column, "companies_column", "Компании группы (ЕРЗ)")
        ),
        "inn_keyword": str(pick(args.inn_keyword, "inn_keyword", "ИНН")),
        "company_separators": list(
            pick(args.company_separators, "company_separators", [";", ","])
        ),
        "multi_value_joiner": str(pick(None, "multi_value_joiner", ";\n")),
        "output_columns": dict(
            pick(
                None,
                "output_columns",
                {
                    "inn": "ИНН",
                    "company_name": "Наименование",
                    "region_name": "Наименование, регион",
                    "region_count": "Кол-во регионов",
                    "region_base_name": "Наименование без города",
                    "region_base_count": "Кол-во наименований без города",
                },
            )
        ),
        "log_to_file": bool(pick(None, "log_to_file", True)),
        "logs_dir": str(pick(None, "logs_dir", "log")),
        "log_file_prefix": str(pick(None, "log_file_prefix", "INFO_base_kg_inn")),
        "output_format": dict(pick(None, "output_format", {})),
    }

    if not settings["input_xlsx"]:
        raise ValueError("Не задан input_xlsx в config.json / CLI")
    return settings


def make_arg_parser() -> argparse.ArgumentParser:
    """CLI аргументы скрипта."""
    parser = argparse.ArgumentParser(
        description="Разбор _base_kg → уникальные ИНН из «Компании группы (ЕРЗ)»"
    )
    parser.add_argument(
        "--config-json",
        default=str(КОНФИГ_ПО_УМОЛЧАНИЮ),
        help="Путь к config.json",
    )
    parser.add_argument("--input-xlsx", default=None, help="Входной Excel")
    parser.add_argument("--output-xlsx", default=None, help="Выходной Excel")
    parser.add_argument(
        "--output-add-timestamp",
        default=None,
        type=lambda v: str(v).lower() in {"1", "true", "yes", "да"},
        help="Добавить таймштамп к имени выходного файла",
    )
    parser.add_argument("--output-sheet", default=None, help="Имя листа результата")
    parser.add_argument("--base-table", default=None, help="Имя смарт-таблицы")
    parser.add_argument("--region-column", default=None, help="Колонка «Наименование, регион»")
    parser.add_argument(
        "--companies-column",
        default=None,
        help="Колонка «Компании группы (ЕРЗ)»",
    )
    parser.add_argument("--inn-keyword", default=None, help="Ключевое слово перед номером ИНН")
    parser.add_argument(
        "--company-separators",
        nargs="+",
        default=None,
        help="Разделители компаний, например: ; ,",
    )
    return parser


def main() -> None:
    """Точка входа: config → чтение → разбор → агрегация → Excel."""
    parser = make_arg_parser()
    args = parser.parse_args()
    settings = разрешить_настройки(args)
    config_dir = Path(str(settings["_config_dir"])).resolve()

    log_path = настроить_логирование(
        logs_dir=разрешить_путь(str(settings["logs_dir"]), config_dir),
        log_file_prefix=str(settings["log_file_prefix"]),
        log_to_file=bool(settings["log_to_file"]),
    )
    if log_path is not None:
        logger.info("Файл лога: %s", log_path)

    input_xlsx = разрешить_путь(str(settings["input_xlsx"]), config_dir)
    output_xlsx = разрешить_путь(str(settings["output_xlsx"]), config_dir)
    if settings["output_add_timestamp"]:
        output_xlsx = с_таймштампом(
            output_xlsx, str(settings["output_timestamp_format"])
        )

    if not input_xlsx.exists():
        raise FileNotFoundError(f"Входной файл не найден: {input_xlsx}")

    source_rows = прочитать_смарт_таблицу(input_xlsx, str(settings["base_table"]))
    aggregated = агрегировать_по_инн(
        rows=source_rows,
        region_column=str(settings["region_column"]),
        companies_column=str(settings["companies_column"]),
        inn_keyword=str(settings["inn_keyword"]),
        separators=list(settings["company_separators"]),
        joiner=str(settings["multi_value_joiner"]),
    )

    сохранить_результат(
        rows=aggregated,
        output_path=output_xlsx,
        sheet_name=str(settings["output_sheet"]),
        output_columns=dict(settings["output_columns"]),
        format_cfg=dict(settings["output_format"]),
    )
    logger.info("Готово.")


if __name__ == "__main__":
    main()
