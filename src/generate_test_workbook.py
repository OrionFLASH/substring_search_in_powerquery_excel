#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Генерация тестовой Excel-книги с полным набором ключевых полей и сценариев."""

from __future__ import annotations

import json
import random
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_CONFIG = ROOT / "config.json"
LEGACY_WORKBOOK = ROOT / "input" / "workbook.xlsx"


BASE_HEADERS: list[str] = [
    "Наименование, регион",
    "key_and_full_1",
    "key_and_full_2",
    "key_and_full_3",
    "key_and_not_1",
    "key_and_not_2",
    "key_and_not_3",
    "key_and_non_1",
    "key_and_non_2",
    "key_or_full_1",
    "key_or_full_2",
    "key_or_full_3",
    "key_or_not_1",
    "key_or_not_2",
    "key_or_not_3",
    "key_or_non_1",
    "key_or_non_2",
    "key_fix_id",
    "Сценарий_ключа",
    "ключи_задублированы",
    "комментарий_ключей",
]

HOLDING_HEADERS: list[str] = [
    "ID холдинга",
    "Холдинг",
    "Sum([ОД текущий год])",
    "Sum([ОД прошлый год])",
    "Сценарий",
    "Целевое_ГСЗ",
    "Ожидаемый_статус",
]


def empty_base_row() -> dict[str, Any]:
    return {h: None for h in BASE_HEADERS}


def base_row(**kwargs: Any) -> dict[str, Any]:
    row = empty_base_row()
    row.update(kwargs)
    return row


def holding_row(**kwargs: Any) -> dict[str, Any]:
    row: dict[str, Any] = {h: None for h in HOLDING_HEADERS}
    row.update(kwargs)
    return row


def scenario_base_rows() -> list[dict[str, Any]]:
    """Строки справочника с разными комбинациями ключей и key_fix_id."""
    p = "zqscn"  # префикс, не встречающийся в bulk-данных
    rows: list[dict[str, Any]] = []

    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND full — тест, Москва",
                "key_and_full_1": f"{p}full",
                "Сценарий_ключа": "S_AND_FULL",
                "комментарий_ключей": "Подстрока full",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND not — тест, Москва",
                "key_and_not_1": f"{p}not",
                "Сценарий_ключа": "S_AND_NOT",
                "комментарий_ключей": "Отдельное слово; «zqscnnotnik» не матчится",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND full+not — тест, Москва",
                "key_and_full_1": f"{p}sam",
                "key_and_not_1": f"{p}samword",
                "Сценарий_ключа": "S_AND_FULL_NOT",
                "комментарий_ключей": "Подстрока + отдельное слово",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} OR full — тест, СПб",
                "key_or_full_1": f"{p}or1",
                "key_or_full_2": f"{p}or2",
                "Сценарий_ключа": "S_OR_FULL",
                "комментарий_ключей": "Хотя бы одна OR-подстрока",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} OR not — тест, СПб",
                "key_or_not_1": f"{p}orn1",
                "key_or_not_2": f"{p}orn2",
                "Сценарий_ключа": "S_OR_NOT",
                "комментарий_ключей": "Хотя бы одно OR-слово",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND non — тест, Москва",
                "key_and_full_1": f"{p}nonok",
                "key_and_non_1": f"{p}excl",
                "Сценарий_ключа": "S_AND_NON",
                "комментарий_ключей": "Исключение при «zqscnexcl» в compact-тексте",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} OR non — тест, Москва",
                "key_and_full_1": f"{p}ornon",
                "key_or_non_1": f"{p}block",
                "Сценарий_ключа": "S_OR_NON",
                "комментарий_ключей": "Исключение при «zqscnblock» в compact-тексте",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND not overlap — тест, Москва",
                "key_and_not_1": f"{p}pig",
                "key_and_not_2": f"{p}needle",
                "Сценарий_ключа": "S_AND_NOT_OVERLAP",
                "комментарий_ключей": "Два слова без пересечения позиций",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} AND non pair — тест, СПб",
                "key_and_non_1": f"{p}vk1",
                "key_and_non_2": f"{p}vk2",
                "key_and_full_1": f"{p}vk1",
                "Сценарий_ключа": "S_AND_NON_PAIR",
                "комментарий_ключей": "Исключение только если оба non найдены",
            }
        )
    )

    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Fix F1 — зафиксированный бренд А",
                "key_fix_id": "100",
                "Сценарий_ключа": "FIX_F1",
                "комментарий_ключей": "Статус: зафиксированное значение",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Fix F2 — fallback на ключ",
                "key_fix_id": "999",
                "key_and_full_1": f"{p}fixfb",
                "Сценарий_ключа": "FIX_F2",
                "комментарий_ключей": "ID 999 нет; ключ срабатывает",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Fix F3 — fallback без ключей",
                "key_fix_id": "999",
                "key_and_full_1": f"{p}fixmiss",
                "Сценарий_ключа": "FIX_F3",
                "комментарий_ключей": "ID 999 нет; ключ не матчится",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Fix partial — бренд B",
                "key_fix_id": "100; 999",
                "Сценарий_ключа": "FIX_PARTIAL",
                "комментарий_ключей": "Часть fix-ID найдена",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Fix multi — бренд C и D",
                "key_fix_id": "100; 200",
                "Сценарий_ключа": "FIX_MULTI",
                "комментарий_ключей": "Два fix-ID на одной строке",
            }
        )
    )

    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Multi A — дубль 1, Москва",
                "key_and_full_1": f"{p}multi",
                "Сценарий_ключа": "MULTIPLE_A",
                "ключи_задублированы": "да",
                "комментарий_ключей": "Два ГСЗ с одним ключом",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Multi B — дубль 2, Москва",
                "key_and_full_1": f"{p}multi",
                "Сценарий_ключа": "MULTIPLE_B",
                "ключи_задублированы": "да",
                "комментарий_ключей": "Два ГСЗ с одним ключом",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} No match — уникальный ключ",
                "key_and_full_1": f"{p}zzzztest",
                "Сценарий_ключа": "NO_MATCH",
                "комментарий_ключей": "Нет подходящего холдинга",
            }
        )
    )
    rows.append(
        base_row(
            **{
                "Наименование, регион": f"{p} Пустые ключи — не участвует",
                "Сценарий_ключа": "EMPTY_KEYS",
                "комментарий_ключей": "Строка без ключей",
            }
        )
    )

    return rows


def scenario_holding_rows() -> list[dict[str, Any]]:
    """Холдинги, сопоставленные со сценарными строками справочника."""
    rnd = random.Random(42)
    rows: list[dict[str, Any]] = []

    def add(
        holding_id: int,
        name: str,
        scenario: str,
        target_gsz: str,
        expected_status: str,
    ) -> None:
        od1 = round(rnd.uniform(100, 20000), 2)
        od2 = round(rnd.uniform(100, 15000), 2)
        rows.append(
            holding_row(
                **{
                    "ID холдинга": holding_id,
                    "Холдинг": name,
                    "Sum([ОД текущий год])": od1,
                    "Sum([ОД прошлый год])": od2,
                    "Сценарий": scenario,
                    "Целевое_ГСЗ": target_gsz,
                    "Ожидаемый_статус": expected_status,
                }
            )
        )

    p = "zqscn"

    add(100, "Fix Hold А — зафиксированный", "FIX_F1", f"{p} Fix F1 — зафиксированный бренд А", "зафиксированное значение")
    add(200, "Fix Hold B — второй fix", "FIX_MULTI", f"{p} Fix multi — бренд C и D", "зафиксированное значение")
    add(101, f"ООО {p}full Group", "S_AND_FULL", f"{p} AND full — тест, Москва", "найдено соответствие")
    add(102, f"ООО {p}notnik", "S_AND_NOT_NEG", "-", "-")
    add(103, f"ГК {p}samx {p}samword", "S_AND_FULL_NOT", f"{p} AND full+not — тест, Москва", "найдено соответствие")
    add(104, f"ООО {p}or1 alpha", "S_OR_FULL", f"{p} OR full — тест, СПб", "найдено соответствие")
    add(105, f"ООО {p}orn1 Group", "S_OR_NOT", f"{p} OR not — тест, СПб", "найдено соответствие")
    add(106, f"ГК {p}nonok {p}excl", "S_AND_NON", "-", "-")
    add(107, f"ГК {p}ornon {p}block", "S_OR_NON", "-", "-")
    add(108, f"ООО {p}pig {p}needle", "S_AND_NOT_OVERLAP", f"{p} AND not overlap — тест, Москва", "найдено соответствие")
    add(109, f"ООО {p}pig{p}needle", "S_AND_NOT_OVERLAP_NEG", "-", "-")
    add(110, f"ООО {p}vk1 {p}vk2", "S_AND_NON_PAIR", "-", "-")
    add(111, f"ООО {p}vk1", "S_AND_NON_PAIR_OK", f"{p} AND non pair — тест, СПб", "найдено соответствие")
    add(112, f"ООО {p}fixfb", "FIX_F2", f"{p} Fix F2 — fallback на ключ", "найдено соответствие")
    add(113, f"АО {p}fixf3_unknown", "FIX_F3", "-", "-")
    add(114, f"ООО {p}multi", "MULTIPLE", f"{p} Multi A — дубль 1, Москва", "есть пересечения по ключам")
    add(115, "ООО ZZZ Unique", "NO_MATCH", "-", "-")
    add(116, f"ООО {p}not x", "S_AND_NOT", f"{p} AND not — тест, Москва", "найдено соответствие")

    return rows


def migrate_legacy_base_rows() -> list[dict[str, Any]]:
    """Перенос массовых строк из старой книги workbook.xlsx."""
    if not LEGACY_WORKBOOK.exists():
        return []

    wb = load_workbook(LEGACY_WORKBOOK, read_only=True, data_only=True)
    ws = wb["_base_gsz"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {str(h): i for i, h in enumerate(headers) if h}

    def val(row: tuple[Any, ...], name: str) -> Any:
        i = idx.get(name)
        if i is None:
            return None
        return row[i] if i < len(row) else None

    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        out.append(
            base_row(
                **{
                    "Наименование, регион": val(row, "Наименование, регион"),
                    "key_and_full_1": val(row, "key_and_full_1"),
                    "key_and_full_2": val(row, "key_and_full_2"),
                    "key_and_full_3": val(row, "key_and_full_3"),
                    "key_and_not_1": val(row, "key_and_not_1"),
                    "key_and_not_2": val(row, "key_and_not_2"),
                    "key_and_not_3": val(row, "key_and_not_3"),
                    "key_or_full_1": val(row, "key_or_full_1"),
                    "key_or_full_2": val(row, "key_or_full_2"),
                    "key_or_full_3": val(row, "key_or_full_3"),
                    "key_or_not_1": val(row, "key_or_not_1"),
                    "key_or_not_2": val(row, "key_or_not_2"),
                    "key_or_not_3": val(row, "key_or_not_3"),
                    "Сценарий_ключа": "BULK",
                    "ключи_задублированы": val(row, "ключи_задублированы"),
                    "комментарий_ключей": val(row, "комментарий_ключей"),
                }
            )
        )
    wb.close()
    return out


def migrate_legacy_holding_rows() -> list[dict[str, Any]]:
    if not LEGACY_WORKBOOK.exists():
        return []

    wb = load_workbook(LEGACY_WORKBOOK, read_only=True, data_only=True)
    ws = wb["_HOLD_OD"]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        out.append(
            holding_row(
                **{
                    "ID холдинга": row[0],
                    "Холдинг": row[1],
                    "Sum([ОД текущий год])": row[2],
                    "Sum([ОД прошлый год])": row[3],
                    "Сценарий": row[4],
                    "Целевое_ГСЗ": row[5],
                    "Ожидаемый_статус": None,
                }
            )
        )
    wb.close()
    return out


def write_sheet(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def add_smart_table(ws: Worksheet, table_name: str) -> None:
    ref = f"A1:{chr(64 + ws.max_column)}{ws.max_row}"
    if ws.max_column > 26:
        from openpyxl.utils import get_column_letter

        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook(output_path: Path) -> dict[str, int]:
    scenario_base = scenario_base_rows()
    scenario_hold = scenario_holding_rows()
    bulk_base = migrate_legacy_base_rows()
    bulk_hold = migrate_legacy_holding_rows()

    base_rows = scenario_base + bulk_base
    hold_rows = scenario_hold + bulk_hold

    wb = Workbook()
    ws_base = wb.active
    ws_base.title = "_base_gsz"
    write_sheet(ws_base, BASE_HEADERS, base_rows)
    add_smart_table(ws_base, "_base_gsz")

    ws_hold = wb.create_sheet("_HOLD_OD")
    write_sheet(ws_hold, HOLDING_HEADERS, hold_rows)
    add_smart_table(ws_hold, "_HOLD_OD")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "scenario_base": len(scenario_base),
        "scenario_holdings": len(scenario_hold),
        "bulk_base": len(bulk_base),
        "bulk_holdings": len(bulk_hold),
        "total_base": len(base_rows),
        "total_holdings": len(hold_rows),
    }


def resolve_output_path() -> Path:
    with DEFAULT_CONFIG.open(encoding="utf-8") as f:
        cfg = json.load(f)
    rel = cfg["gsz_matcher_parallel"]["input_xlsx"]
    return (ROOT / rel).resolve()


def main() -> None:
    output_path = resolve_output_path()
    stats = build_workbook(output_path)
    print(f"Создано: {output_path}")
    for key, value in stats.items():
        print(f"  {key}: {value}")


if __name__ == "__main__":
    main()
