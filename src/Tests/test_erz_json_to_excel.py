#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Тесты преобразования ERZ JSON → строки Excel."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from erz_json_to_excel import (
    построить_лист_групп,
    построить_лист_компаний,
    сохранить_excel,
)

КОРЕНЬ = Path(__file__).resolve().parent.parent.parent
SAMPLE_JSON = КОРЕНЬ / "input" / "erz_full_sample.json"


class TestErzJsonToExcel(unittest.TestCase):
    """Проверки на тестовом erz_full_sample.json."""

    @classmethod
    def setUpClass(cls) -> None:
        with SAMPLE_JSON.open(encoding="utf-8") as f:
            cls.data = json.load(f)
        cls.joiner = ";\n"

    def test_sample_exists(self) -> None:
        self.assertTrue(SAMPLE_JSON.is_file())

    def test_groups_sheet(self) -> None:
        rows = построить_лист_групп(
            self.data,
            joiner=self.joiner,
            company_template="{name} (ИНН {inn})",
        )
        self.assertEqual(len(rows), 2)
        pik = next(r for r in rows if r["group_name"] == "ПИК")
        self.assertIn("г.Москва", pik["regions"])
        self.assertIn("Московская область", pik["regions"])
        self.assertEqual(pik["region_count"], 2)
        self.assertEqual(pik["group_companies_count"], 2)
        self.assertEqual(pik["brand_companies_count"], 2)
        self.assertIn("ПАО ПИК СЗ (ИНН 7713011336)", pik["group_companies"])
        self.assertIn(";\n", pik["regions"])

    def test_companies_roles_and_groups(self) -> None:
        rows = построить_лист_компаний(
            self.data,
            joiner=self.joiner,
            group_region_template="{group}, {region}",
            role_labels={
                "group": "компания группы",
                "brand": "компания бренда",
                "both": "группа и бренд",
            },
        )
        by_inn = {r["inn"]: r for r in rows}
        self.assertEqual(len(by_inn), 5)

        both = by_inn["7701511447"]
        self.assertEqual(both["role"], "группа и бренд")
        self.assertEqual(both["groups_count"], 2)
        self.assertIn("ПИК", both["groups"])
        self.assertIn("Самолёт", both["groups"])
        self.assertIn("ПИК, г.Москва", both["group_regions"])
        self.assertIn("Самолёт, Московская область", both["group_regions"])
        self.assertGreaterEqual(both["group_regions_count"], 2)

        group_only = by_inn["7713011336"]
        self.assertEqual(group_only["role"], "компания группы")
        self.assertEqual(group_only["groups_count"], 1)

        brand_only = by_inn["7729755852"]
        self.assertEqual(brand_only["role"], "компания бренда")
        self.assertEqual(brand_only["groups_count"], 2)

        partner = by_inn["7703009999"]
        self.assertEqual(partner["role"], "компания бренда")
        self.assertEqual(partner["groups_count"], 1)

    def test_save_workbook(self) -> None:
        group_rows = построить_лист_групп(
            self.data, joiner=self.joiner, company_template="{name} (ИНН {inn})"
        )
        company_rows = построить_лист_компаний(
            self.data,
            joiner=self.joiner,
            group_region_template="{group}, {region}",
            role_labels={
                "group": "компания группы",
                "brand": "компания бренда",
                "both": "группа и бренд",
            },
        )
        settings = {
            "groups_sheet": "_ERZ_GROUPS",
            "companies_sheet": "_ERZ_COMPANIES",
            "groups_columns": {
                "group_id": "ID группы",
                "group_name": "Группа застройщиков",
                "group_url_id": "urlId группы",
                "regions": "Регионы",
                "region_count": "Кол-во регионов",
                "group_companies": "Компании группы",
                "group_companies_count": "Кол-во компаний группы",
                "brand_companies": "Компании бренда",
                "brand_companies_count": "Кол-во компаний бренда",
            },
            "companies_columns": {
                "inn": "ИНН",
                "company_name": "Наименование",
                "ogrn": "ОГРН",
                "url_id": "urlId",
                "role": "Тип",
                "groups": "Группы застройщиков",
                "groups_count": "Кол-во групп",
                "group_regions": "Группа, регион",
                "group_regions_count": "Кол-во группа+регион",
            },
            "groups_format": {
                "freeze_rows": 1,
                "freeze_cols": 1,
                "columns": {"group_id": {"width": 18}},
            },
            "companies_format": {
                "freeze_rows": 1,
                "freeze_cols": 1,
                "columns": {"inn": {"width": 20, "align": "center"}},
            },
        }
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "erz_test.xlsx"
            сохранить_excel(out, group_rows, company_rows, settings)
            self.assertTrue(out.is_file())
            from openpyxl import load_workbook

            wb = load_workbook(out)
            self.assertEqual(wb.sheetnames, ["_ERZ_GROUPS", "_ERZ_COMPANIES"])
            self.assertEqual(wb["_ERZ_GROUPS"].max_row, 3)
            self.assertEqual(wb["_ERZ_COMPANIES"].max_row, 6)
            self.assertIsNotNone(wb["_ERZ_GROUPS"].auto_filter.ref)
            self.assertEqual(wb["_ERZ_GROUPS"].freeze_panes, "B2")
            wb.close()


if __name__ == "__main__":
    unittest.main()
