#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Тесты разбора _base_kg → уникальные ИНН."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from extract_base_kg_inn import (
    агрегировать_по_инн,
    имя_до_запятой,
    разобрать_компании_ерз,
    сохранить_результат,
)


class TestParseCompaniesErz(unittest.TestCase):
    """Разбор ячейки «Компании группы (ЕРЗ)»."""

    def test_три_компании_через_точку_с_запятой(self) -> None:
        text = (
            "АО 3-Й ТАКСОМОТОРНЫЙ ПАРК (ИНН 1000000001); "
            "АО ЖБИ-23 (ИНН 1000000002); "
            "АО ЗемПроектСтрой (ИНН 1000000003)"
        )
        pairs = разобрать_компании_ерз(text, "ИНН", [";", ","])
        self.assertEqual(
            pairs,
            [
                ("АО 3-Й ТАКСОМОТОРНЫЙ ПАРК", "1000000001"),
                ("АО ЖБИ-23", "1000000002"),
                ("АО ЗемПроектСтрой", "1000000003"),
            ],
        )

    def test_разделитель_запятая(self) -> None:
        text = "ООО Альфа (ИНН 111), ООО Бета (ИНН 222)"
        pairs = разобрать_компании_ерз(text, "ИНН", [";", ","])
        self.assertEqual(pairs, [("ООО Альфа", "111"), ("ООО Бета", "222")])

    def test_ключевое_слово_без_учёта_регистра(self) -> None:
        text = "ООО Тест (инн 999)"
        pairs = разобрать_компании_ерз(text, "ИНН", [";"])
        self.assertEqual(pairs, [("ООО Тест", "999")])

    def test_пустое_и_мусор(self) -> None:
        self.assertEqual(разобрать_компании_ерз(None, "ИНН", [";"]), [])
        self.assertEqual(разобрать_компании_ерз("", "ИНН", [";"]), [])
        self.assertEqual(разобрать_компании_ерз("без скобок", "ИНН", [";"]), [])


class TestNameBeforeComma(unittest.TestCase):
    """Урезание «Наименование, регион» до части до запятой."""

    def test_с_городом(self) -> None:
        self.assertEqual(имя_до_запятой("ПИК, Москва"), "ПИК")
        self.assertEqual(имя_до_запятой("ПИК, Иваново"), "ПИК")

    def test_без_запятой(self) -> None:
        self.assertEqual(имя_до_запятой("Самолет"), "Самолет")

    def test_пусто(self) -> None:
        self.assertEqual(имя_до_запятой(""), "")
        self.assertEqual(имя_до_запятой("  , Москва"), "")


class TestAggregateByInn(unittest.TestCase):
    """Агрегация уникальных ИНН."""

    def test_несколько_имён_и_регионов(self) -> None:
        rows = [
            {
                "Наименование, регион": "Группа А, Москва",
                "Компании группы (ЕРЗ)": (
                    "АО Альфа (ИНН 100); ООО Альфа Трейд (ИНН 100)"
                ),
            },
            {
                "Наименование, регион": "Группа А, СПб",
                "Компании группы (ЕРЗ)": "АО Альфа (ИНН 100)",
            },
            {
                "Наименование, регион": "Группа Б, Казань",
                "Компании группы (ЕРЗ)": "АО Бета (ИНН 200)",
            },
        ]
        out = агрегировать_по_инн(
            rows=rows,
            region_column="Наименование, регион",
            companies_column="Компании группы (ЕРЗ)",
            inn_keyword="ИНН",
            separators=[";", ","],
            joiner=";\n",
        )
        self.assertEqual(len(out), 2)
        by_inn = {r["inn"]: r for r in out}

        self.assertIn("АО Альфа", by_inn["100"]["company_name"])
        self.assertIn("ООО Альфа Трейд", by_inn["100"]["company_name"])
        self.assertIn(";\n", by_inn["100"]["company_name"])
        self.assertIn("Группа А, Москва", by_inn["100"]["region_name"])
        self.assertIn("Группа А, СПб", by_inn["100"]["region_name"])
        self.assertEqual(by_inn["100"]["region_count"], 2)
        self.assertEqual(by_inn["100"]["region_base_name"], "Группа А")
        self.assertEqual(by_inn["100"]["region_base_count"], 1)

        self.assertEqual(by_inn["200"]["company_name"], "АО Бета")
        self.assertEqual(by_inn["200"]["region_name"], "Группа Б, Казань")
        self.assertEqual(by_inn["200"]["region_count"], 1)
        self.assertEqual(by_inn["200"]["region_base_name"], "Группа Б")
        self.assertEqual(by_inn["200"]["region_base_count"], 1)

    def test_разные_имена_без_города(self) -> None:
        rows = [
            {
                "Наименование, регион": "ПИК, Москва",
                "Компании группы (ЕРЗ)": "ООО П (ИНН 1)",
            },
            {
                "Наименование, регион": "ПИК, Иваново",
                "Компании группы (ЕРЗ)": "ООО П (ИНН 1)",
            },
            {
                "Наименование, регион": "Самолет, СПб",
                "Компании группы (ЕРЗ)": "ООО П (ИНН 1)",
            },
        ]
        out = агрегировать_по_инн(
            rows=rows,
            region_column="Наименование, регион",
            companies_column="Компании группы (ЕРЗ)",
            inn_keyword="ИНН",
            separators=[";"],
            joiner=";\n",
        )
        row = out[0]
        self.assertEqual(row["region_count"], 3)
        self.assertEqual(row["region_base_count"], 2)
        self.assertIn("ПИК", row["region_base_name"])
        self.assertIn("Самолет", row["region_base_name"])
        self.assertIn(";\n", row["region_base_name"])


class TestExcelOutputFormatting(unittest.TestCase):
    """Проверка форматирования выходного Excel."""

    def test_форматирование_и_значения(self) -> None:
        rows = [
            {
                "inn": "100",
                "company_name": "АО Альфа;\nООО Альфа Трейд",
                "region_name": "Группа А, Москва;\nГруппа А, СПб",
                "region_count": 2,
                "region_base_name": "Группа А",
                "region_base_count": 1,
            }
        ]
        format_cfg = {
            "header_bold": True,
            "header_center": True,
            "header_wrap": True,
            "freeze_rows": 1,
            "freeze_cols": 1,
            "data_vertical_center": True,
            "data_horizontal": "left",
            "columns": {
                "inn": {"width": 30, "wrap": False},
                "company_name": {"width": 70, "wrap": True},
                "region_name": {"width": 100, "wrap": True},
                "region_count": {"width": 10, "wrap": False, "align": "center"},
                "region_base_name": {"width": 90, "wrap": True},
                "region_base_count": {"width": 10, "wrap": False, "align": "center"},
            },
        }
        output_columns = {
            "inn": "ИНН",
            "company_name": "Наименование",
            "region_name": "Наименование, регион",
            "region_count": "Кол-во регионов",
            "region_base_name": "Наименование без города",
            "region_base_count": "Кол-во наименований без города",
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "out.xlsx"
            сохранить_результат(
                rows=rows,
                output_path=path,
                sheet_name="_BASE_KG_INN",
                output_columns=output_columns,
                format_cfg=format_cfg,
            )
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(ws["A1"].value, "ИНН")
            self.assertEqual(ws["B1"].value, "Наименование")
            self.assertEqual(ws["C1"].value, "Наименование, регион")
            self.assertEqual(ws["D1"].value, "Кол-во регионов")
            self.assertEqual(ws["E1"].value, "Наименование без города")
            self.assertEqual(ws["F1"].value, "Кол-во наименований без города")
            self.assertTrue(ws["A1"].font.bold)
            self.assertEqual(ws["A1"].alignment.horizontal, "center")
            self.assertEqual(ws["A1"].alignment.vertical, "center")
            self.assertTrue(ws["A1"].alignment.wrap_text)

            self.assertEqual(ws["A2"].value, "100")
            self.assertEqual(ws["A2"].alignment.horizontal, "left")
            self.assertEqual(ws["A2"].alignment.vertical, "center")
            self.assertFalse(bool(ws["A2"].alignment.wrap_text))
            self.assertTrue(ws["B2"].alignment.wrap_text)

            self.assertEqual(ws["D2"].value, 2)
            self.assertEqual(ws["D2"].alignment.horizontal, "center")
            self.assertEqual(ws["D2"].alignment.vertical, "center")
            self.assertEqual(ws["E2"].value, "Группа А")
            self.assertEqual(ws["F2"].value, 1)
            self.assertEqual(ws["F2"].alignment.horizontal, "center")

            self.assertEqual(ws.column_dimensions["A"].width, 30)
            self.assertEqual(ws.column_dimensions["B"].width, 70)
            self.assertEqual(ws.column_dimensions["C"].width, 100)
            self.assertEqual(ws.column_dimensions["D"].width, 10)
            self.assertEqual(ws.column_dimensions["E"].width, 90)
            self.assertEqual(ws.column_dimensions["F"].width, 10)
            self.assertEqual(ws.freeze_panes, "B2")
            self.assertIsNotNone(ws.auto_filter.ref)
            wb.close()


class TestEndToEndSampleWorkbook(unittest.TestCase):
    """Сквозной прогон на временной книге с таблицей _base_kg."""

    def test_read_aggregate_via_temp_workbook(self) -> None:
        from extract_base_kg_inn import прочитать_смарт_таблицу

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "base_kg.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "_base_kg"
            ws.append(["Наименование, регион", "Компании группы (ЕРЗ)"])
            ws.append(
                [
                    "Таксопарк, Москва",
                    "АО 3-Й ТАКСОМОТОРНЫЙ ПАРК (ИНН 1000000001); "
                    "АО ЖБИ-23 (ИНН 1000000002)",
                ]
            )
            ws.append(
                [
                    "ЖБИ, Казань",
                    "АО ЖБИ-23 (ИНН 1000000002); АО Новый (ИНН 1000000002)",
                ]
            )
            table = Table(displayName="_base_kg", ref="A1:B3")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
            ws.add_table(table)
            wb.save(path)
            wb.close()

            source = прочитать_смарт_таблицу(path, "_base_kg")
            self.assertEqual(len(source), 2)
            out = агрегировать_по_инн(
                rows=source,
                region_column="Наименование, регион",
                companies_column="Компании группы (ЕРЗ)",
                inn_keyword="ИНН",
                separators=[";", ","],
                joiner=";\n",
            )
            by_inn = {r["inn"]: r for r in out}
            self.assertEqual(set(by_inn), {"1000000001", "1000000002"})
            self.assertEqual(by_inn["1000000001"]["company_name"], "АО 3-Й ТАКСОМОТОРНЫЙ ПАРК")
            self.assertEqual(by_inn["1000000001"]["region_count"], 1)
            self.assertEqual(by_inn["1000000001"]["region_base_name"], "Таксопарк")
            self.assertIn("АО ЖБИ-23", by_inn["1000000002"]["company_name"])
            self.assertIn("АО Новый", by_inn["1000000002"]["company_name"])
            self.assertEqual(by_inn["1000000002"]["region_count"], 2)
            self.assertEqual(by_inn["1000000002"]["region_base_count"], 2)
            self.assertIn("Таксопарк", by_inn["1000000002"]["region_base_name"])
            self.assertIn("ЖБИ", by_inn["1000000002"]["region_base_name"])
            self.assertIn("Таксопарк, Москва", by_inn["1000000002"]["region_name"])
            self.assertIn("ЖБИ, Казань", by_inn["1000000002"]["region_name"])


if __name__ == "__main__":
    unittest.main()
