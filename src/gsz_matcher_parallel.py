#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Параллельное сопоставление холдингов с условными ГСЗ.

Скрипт переносит логику Power Query в Python и формирует выходной Excel.

Основной поток:
1. Чтение смарт-таблиц `_HOLD_OD` и `_base_gsz`.
2. Подготовка метаданных справочника (ключи, якорь, non-исключения).
3. Параллельное сопоставление каждого холдинга с кандидатами из справочника.
4. Запись результата на два листа с настраиваемыми колонками.

Группы ключей в `_base_gsz`:
- and_full / and_not — обязательные совпадения (подстрока / отдельное слово);
- or_full / or_not — альтернативные совпадения;
- and_non / or_non — исключения по тексту холдинга без пробелов.

Оптимизации:
- предобработка справочника один раз;
- якорный предфильтр кандидатов;
- кэш позиций токенов в рамках одного холдинга;
- ProcessPoolExecutor для параллельной обработки.
"""

from __future__ import annotations

import argparse
import json
import multiprocessing as mp
import os
import re
import sys
import time
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
from concurrent.futures import wait
from concurrent.futures import FIRST_COMPLETED
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# =============================================================================
# Константы: группы колонок ключей и настройки по умолчанию
# Списки можно переопределить в config.json (and_full_cols, and_non_cols и т.д.)
# =============================================================================
DEFAULT_AND_FULL = ["key_and_full_1", "key_and_full_2", "key_and_full_3"]
DEFAULT_AND_NOT = ["key_and_not_1", "key_and_not_2", "key_and_not_3"]
DEFAULT_AND_NON = ["key_and_non_1", "key_and_non_2"]
DEFAULT_OR_FULL = ["key_or_full_1", "key_or_full_2", "key_or_full_3"]
DEFAULT_OR_NOT = ["key_or_not_1", "key_or_not_2", "key_or_not_3"]
DEFAULT_OR_NON = ["key_or_non_1", "key_or_non_2"]
DEFAULT_FIX_ID_COL = "key_fix_id"
DEFAULT_HOLDING_ID_COLUMN = "ID холдинга"
DEFAULT_MIN_WIDTH_ALL = 30.0
DEFAULT_CONFIG_PATH = Path(__file__).resolve().parent.parent / "config.json"

# Тексты колонки «статус» по умолчанию (переопределяются через match_status_texts в config).
DEFAULT_STATUS_NONE = "-"
DEFAULT_STATUS_SINGLE = "найдено соответствие"
DEFAULT_STATUS_FIXED = "зафиксированное значение"
DEFAULT_STATUS_MULTIPLE = "есть пересечения по ключам"
DEFAULT_STATUS_FIX_NOT_FOUND = "фикс значение не найдено"
DEFAULT_STATUS_FIX_PARTIAL = "найдена часть фиксированных значений"
DEFAULT_STATUS_MULTIPLE_PLACEHOLDER = ":=>"

# Обратная совместимость для тестов и внешних импортов.
STATUS_NONE = DEFAULT_STATUS_NONE
STATUS_SINGLE = DEFAULT_STATUS_SINGLE
STATUS_FIXED = DEFAULT_STATUS_FIXED
STATUS_MULTIPLE = DEFAULT_STATUS_MULTIPLE
STATUS_FIX_NOT_FOUND = DEFAULT_STATUS_FIX_NOT_FOUND
STATUS_FIX_PARTIAL = DEFAULT_STATUS_FIX_PARTIAL
STATUS_MULTIPLE_PLACEHOLDER = DEFAULT_STATUS_MULTIPLE_PLACEHOLDER


@dataclass(frozen=True)
class OutputColumnSpec:
    """Описание добавляемой колонки на выходном листе."""

    key: str
    name: str
    width: float
    wrap: bool = False


DEFAULT_HOLDING_OUTPUT_COLUMNS: tuple[OutputColumnSpec, ...] = (
    OutputColumnSpec("gsz_primary", "условное ГСЗ", 150),
    OutputColumnSpec("gsz_debug", "Отладка_совпадения_ГСЗ", 100, wrap=True),
    OutputColumnSpec("match_status", "статус", 40),
    OutputColumnSpec("match_count", "Кол-во совпадений", 30),
)

DEFAULT_BASE_OUTPUT_COLUMNS: tuple[OutputColumnSpec, ...] = (
    OutputColumnSpec("holding_count", "кол-во холдингов", 30),
    OutputColumnSpec("found_holding", "найденный холдинг", 150),
    OutputColumnSpec("found_holding_debug", "Отладка_найденного_холдинга", 100, wrap=True),
    OutputColumnSpec("match_status", "статус", 40),
    OutputColumnSpec("key_string", "строка ключа", 30),
    OutputColumnSpec("key_length", "длина ключа", 30),
    OutputColumnSpec("key_repeat_count", "число повторов", 30),
)

HOLDING_OUTPUT_COLUMN_KEYS: tuple[str, ...] = tuple(column.key for column in DEFAULT_HOLDING_OUTPUT_COLUMNS)
BASE_OUTPUT_COLUMN_KEYS: tuple[str, ...] = tuple(column.key for column in DEFAULT_BASE_OUTPUT_COLUMNS)


# =============================================================================
# Базовые текстовые операции для режимов full / not / non
# =============================================================================
def normalize_text(value: Any) -> str:
    """Trim + нижний регистр. Используется для холдинга и обычных ключей."""
    if value is None:
        return ""
    return str(value).strip().lower()


def is_letter(ch: str) -> bool:
    """Проверка «буквенности» символа для границ слова в режиме not (латиница + кириллица)."""
    if len(ch) != 1:
        return False
    c = ch.lower()
    return ("a" <= c <= "z") or ("а" <= c <= "я") or c == "ё"


def all_positions_full(text: str, word: str) -> list[tuple[int, int]]:
    """Все вхождения подстроки (режим full). Интервал: (start, end) включительно."""
    if not word or len(word) > len(text):
        return []
    out: list[tuple[int, int]] = []
    start = 0
    wl = len(word)
    while True:
        pos = text.find(word, start)
        if pos == -1:
            break
        out.append((pos, pos + wl - 1))
        start = pos + 1
    return out


def positions_not(text: str, word: str) -> list[tuple[int, int]]:
    """Позиции отдельного слова (режим not): границы — не буква."""
    result: list[tuple[int, int]] = []
    n = len(text)
    for s, e in all_positions_full(text, word):
        left_ok = s == 0 or not is_letter(text[s - 1])
        right_ok = e == n - 1 or not is_letter(text[e + 1])
        if left_ok and right_ok:
            result.append((s, e))
    return result


def overlap(a: tuple[int, int], b: tuple[int, int]) -> bool:
    """Пересекаются ли два интервала вхождений (включительные границы)."""
    return not (a[1] < b[0] or b[1] < a[0])


def and_non_overlapping(position_lists: list[list[tuple[int, int]]], idx: int = 0, chosen: list[tuple[int, int]] | None = None) -> bool:
    """Проверка AND-блока: все токены найдены и их интервалы не пересекаются."""
    chosen = chosen or []
    if idx >= len(position_lists):
        return True
    for interval in position_lists[idx]:
        if any(overlap(interval, prev) for prev in chosen):
            continue
        if and_non_overlapping(position_lists, idx + 1, chosen + [interval]):
            return True
    return False


def extract_words(text: str) -> set[str]:
    # Для fast-предфильтра по режиму "not".
    return set(re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", text.lower()))


@dataclass(frozen=True)
class MatchStatusTexts:
    """Настраиваемые тексты статусов и плейсхолдер для множественных совпадений."""

    none: str = DEFAULT_STATUS_NONE
    single: str = DEFAULT_STATUS_SINGLE
    fixed: str = DEFAULT_STATUS_FIXED
    multiple: str = DEFAULT_STATUS_MULTIPLE
    fix_not_found: str = DEFAULT_STATUS_FIX_NOT_FOUND
    fix_partial: str = DEFAULT_STATUS_FIX_PARTIAL
    multiple_placeholder: str = DEFAULT_STATUS_MULTIPLE_PLACEHOLDER


DEFAULT_MATCH_STATUS_TEXTS = MatchStatusTexts()


@dataclass(frozen=True)
class Token:
    """Один ключ поиска: слово и режим full (подстрока) / not (отдельное слово)."""

    word: str
    is_full: bool


@dataclass(frozen=True)
class BaseMeta:
    """Предобработанная строка справочника _base_gsz для быстрого сопоставления."""

    gsz_value: str
    has_keys: bool
    and_tokens: tuple[Token, ...]
    or_tokens: tuple[Token, ...]
    and_non_tokens: tuple[str, ...]  # исключение AND: все найдены → строка отклоняется
    or_non_tokens: tuple[str, ...]   # исключение OR: любой найден → строка отклоняется
    anchor: Token | None
    fix_ids: tuple[str, ...] = ()  # ID холдингов из key_fix_id
    fix_mode: str = "none"  # none | resolved | fallback


@dataclass(frozen=True)
class SingleHoldingMatchResult:
    """Результат сопоставления одного холдинга со справочником."""

    primary: str
    debug: str
    status: str
    count: int
    matched_indices: tuple[int, ...]
    fixed_indices: tuple[int, ...]


def normalize_holding_id(value: Any) -> str:
    """Нормализация ID холдинга для сравнения (trim, строка)."""
    if value is None:
        return ""
    return str(value).strip()


def parse_fix_ids(value: Any) -> tuple[str, ...]:
    """Разбор key_fix_id: один ID или несколько через «; » / «;»."""
    if value is None:
        return ()
    text = str(value).strip()
    if not text:
        return ()
    parts = re.split(r";\s*", text)
    return tuple(p.strip() for p in parts if p.strip())


def load_match_status_texts(block: dict[str, Any]) -> MatchStatusTexts:
    """Загрузка текстов статусов из config.json (блок gsz_matcher_parallel)."""
    raw = block.get("match_status_texts", {})
    if not isinstance(raw, dict):
        return DEFAULT_MATCH_STATUS_TEXTS
    return MatchStatusTexts(
        none=str(raw.get("none", DEFAULT_STATUS_NONE)),
        single=str(raw.get("single", DEFAULT_STATUS_SINGLE)),
        fixed=str(raw.get("fixed", DEFAULT_STATUS_FIXED)),
        multiple=str(raw.get("multiple", DEFAULT_STATUS_MULTIPLE)),
        fix_not_found=str(raw.get("fix_not_found", DEFAULT_STATUS_FIX_NOT_FOUND)),
        fix_partial=str(raw.get("fix_partial", DEFAULT_STATUS_FIX_PARTIAL)),
        multiple_placeholder=str(raw.get("multiple_placeholder", DEFAULT_STATUS_MULTIPLE_PLACEHOLDER)),
    )


def join_status_lines(lines: list[str]) -> str:
    """Объединение статусов в одну ячейку (каждый с новой строки, без дублей)."""
    unique: list[str] = []
    seen: set[str] = set()
    for line in lines:
        if line and line not in seen:
            seen.add(line)
            unique.append(line)
    return "\n".join(unique) if unique else ""


def resolve_fix_mode(fix_ids: tuple[str, ...], holdings_id_set: frozenset[str]) -> str:
    """Режим fix: none | resolved (все ID) | partial (часть ID) | fallback (ни одного)."""
    if not fix_ids:
        return "none"
    found_count = sum(1 for fid in fix_ids if fid in holdings_id_set)
    if found_count == 0:
        return "fallback"
    if found_count == len(fix_ids):
        return "resolved"
    return "partial"


def compute_holding_status_lines(
    match_count: int,
    fixed_count: int,
    texts: MatchStatusTexts,
) -> str:
    """Статусы для листа холдингов — все применимые, каждый с новой строки."""
    if match_count == 0:
        return texts.none
    lines: list[str] = []
    key_count = match_count - fixed_count
    if fixed_count > 0:
        lines.append(texts.fixed)
    if key_count > 1:
        lines.append(texts.multiple)
    elif key_count == 1 and fixed_count == 0:
        lines.append(texts.single)
    elif key_count == 1 and fixed_count > 0:
        lines.append(texts.single)
    return join_status_lines(lines) or texts.none


def compute_base_row_status_lines(
    meta: BaseMeta,
    matched_pairs: list[tuple[int, bool]],
    texts: MatchStatusTexts,
) -> str:
    """Статусы для строки _base_gsz — накопление всех применимых сообщений."""
    fixed_pairs = [pair for pair in matched_pairs if pair[1]]
    key_pairs = [pair for pair in matched_pairs if not pair[1]]
    lines: list[str] = []

    if meta.fix_ids:
        if meta.fix_mode == "fallback":
            lines.append(texts.fix_not_found)
        elif meta.fix_mode == "partial":
            lines.append(texts.fix_partial)
        elif meta.fix_mode == "resolved" and fixed_pairs:
            lines.append(texts.fixed)

    if len(key_pairs) > 1:
        lines.append(texts.multiple)
    elif len(key_pairs) == 1:
        lines.append(texts.single)

    if not lines and not matched_pairs:
        return texts.fix_not_found if meta.fix_mode == "fallback" else texts.none

    return join_status_lines(lines) or texts.none


def compute_found_holding_primary(
    fixed_entries: list[str],
    key_entries: list[str],
    texts: MatchStatusTexts,
) -> str:
    """Колонка «найденный холдинг»: значение, одно совпадение или «:=>»."""
    if fixed_entries:
        if len(fixed_entries) == 1:
            return fixed_entries[0]
        return texts.multiple_placeholder
    if len(key_entries) == 1:
        return key_entries[0]
    if len(key_entries) > 1:
        return texts.multiple_placeholder
    return texts.none


def format_match_columns(
    values: list[str],
    texts: MatchStatusTexts | None = None,
    debug_sep: str = ";\n",
) -> tuple[str, str]:
    """Основная и отладочная колонки: фактические значения, без статусных фраз."""
    empty = (texts.none if texts else DEFAULT_STATUS_NONE)
    if not values:
        return empty, empty
    primary = values[0]
    debug_text = debug_sep.join(values)
    return primary, debug_text


def parse_cols(value: str) -> list[str]:
    """Разбор списка колонок из config.json или CLI (список / строка через запятую)."""
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [v.strip() for v in str(value).split(",") if v.strip()]


def normalize_non_text(value: Any) -> str:
    """Нормализация текста для non-исключений: без пробельных символов."""
    return "".join(ch for ch in normalize_text(value) if not ch.isspace())


def pick_best_anchor(tokens: list[Token]) -> Token | None:
    """Самый длинный токен из группы — лучший якорь для предфильтра кандидатов."""
    if not tokens:
        return None
    return max(tokens, key=lambda t: len(t.word))


def pick_anchor(and_tokens: list[Token], or_tokens: list[Token]) -> Token | None:
    """Якорный токен для предфильтра: приоритет not, затем самый длинный."""
    and_not = [t for t in and_tokens if not t.is_full]
    and_full = [t for t in and_tokens if t.is_full]
    or_not = [t for t in or_tokens if not t.is_full]
    or_full = [t for t in or_tokens if t.is_full]
    return (
        pick_best_anchor(and_not)
        or pick_best_anchor(and_full)
        or pick_best_anchor(or_not)
        or pick_best_anchor(or_full)
    )


def build_meta_row(
    row: dict[str, Any],
    gsz_col: str,
    and_full_cols: list[str],
    and_not_cols: list[str],
    and_non_cols: list[str],
    or_full_cols: list[str],
    or_not_cols: list[str],
    or_non_cols: list[str],
    fix_id_col: str = DEFAULT_FIX_ID_COL,
    holdings_id_set: frozenset[str] | None = None,
) -> BaseMeta:
    """Сборка метаданных одной строки справочника из шести групп ключей и key_fix_id."""
    and_full_raw = [normalize_text(row.get(c, "")) for c in and_full_cols]
    and_not_raw = [normalize_text(row.get(c, "")) for c in and_not_cols]
    and_non_raw = [normalize_non_text(row.get(c, "")) for c in and_non_cols]
    or_full_raw = [normalize_text(row.get(c, "")) for c in or_full_cols]
    or_not_raw = [normalize_text(row.get(c, "")) for c in or_not_cols]
    or_non_raw = [normalize_non_text(row.get(c, "")) for c in or_non_cols]
    fix_ids = parse_fix_ids(row.get(fix_id_col, ""))
    holdings_ids = holdings_id_set if holdings_id_set is not None else frozenset()
    fix_mode = resolve_fix_mode(fix_ids, holdings_ids)

    and_tokens = [Token(w, True) for w in and_full_raw if w] + [Token(w, False) for w in and_not_raw if w]
    or_tokens = [Token(w, True) for w in or_full_raw if w] + [Token(w, False) for w in or_not_raw if w]
    and_non_tokens = tuple(w for w in and_non_raw if w)
    or_non_tokens = tuple(w for w in or_non_raw if w)
    has_keys = bool(and_tokens or or_tokens or and_non_tokens or or_non_tokens)
    anchor = pick_anchor(and_tokens, or_tokens)

    gsz_value = str(row.get(gsz_col, "") or "").strip()
    return BaseMeta(
        gsz_value=gsz_value,
        has_keys=has_keys,
        and_tokens=tuple(and_tokens),
        or_tokens=tuple(or_tokens),
        and_non_tokens=and_non_tokens,
        or_non_tokens=or_non_tokens,
        anchor=anchor,
        fix_ids=fix_ids,
        fix_mode=fix_mode,
    )


def row_matches(text: str, meta: BaseMeta) -> bool:
    """Проверка одной строки справочника: AND → OR → исключения non."""
    text = normalize_text(text)
    if not text or not meta.has_keys:
        return False

    # AND
    if meta.and_tokens:
        position_lists: list[list[tuple[int, int]]] = []
        for t in meta.and_tokens:
            pos = all_positions_full(text, t.word) if t.is_full else positions_not(text, t.word)
            if not pos:
                return False
            position_lists.append(pos)
        if not and_non_overlapping(position_lists):
            return False

    # OR
    if meta.or_tokens:
        ok_or = False
        for t in meta.or_tokens:
            pos = all_positions_full(text, t.word) if t.is_full else positions_not(text, t.word)
            if pos:
                ok_or = True
                break
        if not ok_or:
            return False

    # --- Исключения non: текст холдинга без пробелов ---
    compact_text = normalize_non_text(text)
    if meta.and_non_tokens:
        # AND NON: отклоняем, только если найдены ВСЕ non-ключи без пересечений
        non_and_positions: list[list[tuple[int, int]]] = []
        for token in meta.and_non_tokens:
            pos = all_positions_full(compact_text, token)
            if not pos:
                non_and_positions = []
                break
            non_and_positions.append(pos)
        if non_and_positions and and_non_overlapping(non_and_positions):
            return False

    if meta.or_non_tokens:
        # OR NON: отклоняем, если найден ХОТЯ БЫ ОДИН non-ключ
        for token in meta.or_non_tokens:
            if all_positions_full(compact_text, token):
                return False

    return True


# =============================================================================
# Чтение Excel: смарт-таблицы (ListObject) → list[dict]
# =============================================================================
def find_table_ref(path: Path, table_name: str) -> tuple[str, tuple[int, int, int, int]]:
    """Найти лист и границы смарт-таблицы."""
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries

    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            if table_name in ws.tables:
                min_col, min_row, max_col, max_row = range_boundaries(ws.tables[table_name].ref)
                return ws.title, (min_col, min_row, max_col, max_row)
    finally:
        wb.close()
    raise ValueError(f"Таблица '{table_name}' не найдена в {path}")


def read_excel_table(
    path: Path,
    table_name: str,
    log_enabled: bool = False,
    progress_every_read_rows: int = 1000,
) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    sheet_title, (min_col, min_row, max_col, max_row) = find_table_ref(path, table_name)
    total_rows = max(0, max_row - min_row + 1)
    rows: list[list[Any]] = []

    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_title]
        for idx, row in enumerate(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            ),
            start=1,
        ):
            rows.append(list(row))
            if log_enabled and (idx % max(1, progress_every_read_rows) == 0 or idx == total_rows):
                log(f"[progress-read] {table_name}: read_rows={idx}/{total_rows}")
    finally:
        wb.close()

    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    body = rows[1:]
    total_body = len(body)
    out: list[dict[str, Any]] = []
    for idx, row in enumerate(body, start=1):
        rec = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        out.append(rec)
        if log_enabled and (idx % max(1, progress_every_read_rows) == 0 or idx == total_body):
            log(f"[progress-read] {table_name}: map_rows={idx}/{total_body}")
    return out


def inspect_workbook_objects(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=False)
    table_names: list[str] = []
    for ws in wb.worksheets:
        table_names.extend(list(ws.tables.keys()))

    defined_names: list[str] = []
    try:
        for dn in wb.defined_names.definedName:
            if getattr(dn, "name", None):
                defined_names.append(str(dn.name))
    except Exception:
        # Безопасный fallback для разных версий openpyxl.
        pass

    return {
        "table_names": sorted(set(table_names)),
        "defined_names": sorted(set(defined_names)),
    }


# =============================================================================
# Глобальное состояние воркеров (инициализируется один раз на процесс)
# =============================================================================
BASE_METAS: tuple[BaseMeta, ...] = ()
FIX_INDICES_BY_HOLDING_ID: dict[str, tuple[int, ...]] = {}
ANCHOR_NOT_INDEX: dict[str, tuple[int, ...]] = {}
ANCHOR_FULL_INDEX: dict[str, tuple[int, ...]] = {}
FULL_ANCHOR_WORDS_BY_CH: dict[str, tuple[str, ...]] = {}
NOT_ANCHOR_WORDS_BY_CH: dict[str, tuple[str, ...]] = {}
NO_ANCHOR_INDICES: tuple[int, ...] = ()
MATCH_STATUS_TEXTS: MatchStatusTexts = DEFAULT_MATCH_STATUS_TEXTS
LOG_FILE_PATH: Path | None = None


def worker_init(
    base_metas: tuple[BaseMeta, ...],
    status_texts: MatchStatusTexts | None = None,
) -> None:
    """Инициализация процесса: метаданные справочника + индексы якорных токенов и fix-ID."""
    global BASE_METAS
    global FIX_INDICES_BY_HOLDING_ID
    global MATCH_STATUS_TEXTS
    global ANCHOR_NOT_INDEX
    global ANCHOR_FULL_INDEX
    global FULL_ANCHOR_WORDS_BY_CH
    global NOT_ANCHOR_WORDS_BY_CH
    global NO_ANCHOR_INDICES

    BASE_METAS = base_metas
    if status_texts is not None:
        MATCH_STATUS_TEXTS = status_texts
    fix_index: dict[str, list[int]] = {}
    not_index: dict[str, list[int]] = {}
    full_index: dict[str, list[int]] = {}
    no_anchor: list[int] = []

    for idx, meta in enumerate(base_metas):
        if meta.fix_mode in ("resolved", "partial"):
            for fid in meta.fix_ids:
                fix_index.setdefault(fid, []).append(idx)
        a = meta.anchor
        if a is None:
            no_anchor.append(idx)
            continue
        if a.is_full:
            full_index.setdefault(a.word, []).append(idx)
        else:
            not_index.setdefault(a.word, []).append(idx)

    FIX_INDICES_BY_HOLDING_ID = {k: tuple(v) for k, v in fix_index.items()}

    ANCHOR_NOT_INDEX = {k: tuple(v) for k, v in not_index.items()}
    ANCHOR_FULL_INDEX = {k: tuple(v) for k, v in full_index.items()}
    NO_ANCHOR_INDICES = tuple(no_anchor)

    full_by_ch: dict[str, list[str]] = {}
    for word in ANCHOR_FULL_INDEX:
        for ch in set(word):
            full_by_ch.setdefault(ch, []).append(word)
    FULL_ANCHOR_WORDS_BY_CH = {k: tuple(v) for k, v in full_by_ch.items()}

    not_by_ch: dict[str, list[str]] = {}
    for word in ANCHOR_NOT_INDEX:
        for ch in set(word):
            not_by_ch.setdefault(ch, []).append(word)
    NOT_ANCHOR_WORDS_BY_CH = {k: tuple(v) for k, v in not_by_ch.items()}


def candidate_indices_for_text(text: str) -> list[int]:
    """Индексы кандидатов _base_gsz по якорным ключам (предфильтр)."""
    if not BASE_METAS:
        return []

    mark = bytearray(len(BASE_METAS))
    for idx in NO_ANCHOR_INDICES:
        mark[idx] = 1

    # not-якоря: границы слова, а не токены extract_words (иначе теряются совпадения вроде «мега» в «мега35»)
    seen_not_words: set[str] = set()
    for ch in set(text):
        for w in NOT_ANCHOR_WORDS_BY_CH.get(ch, ()):
            if w in seen_not_words:
                continue
            seen_not_words.add(w)
            if w in text and positions_not(text, w):
                for idx in ANCHOR_NOT_INDEX.get(w, ()):
                    mark[idx] = 1

    seen_full_words: set[str] = set()
    for ch in set(text):
        for w in FULL_ANCHOR_WORDS_BY_CH.get(ch, ()):
            if w in seen_full_words:
                continue
            seen_full_words.add(w)
            if w in text:
                for idx in ANCHOR_FULL_INDEX.get(w, ()):
                    mark[idx] = 1

    return [i for i, v in enumerate(mark) if v]


# =============================================================================
# Сопоставление одного холдинга со справочником
# =============================================================================
def _meta_matches_with_cache(
    text: str,
    compact_text: str,
    meta: BaseMeta,
    get_positions: Any,
) -> bool:
    """Проверка строки справочника по ключам с кэшем позиций (для match_single_holding)."""
    if not text or not meta.has_keys:
        return False

    if meta.and_tokens:
        position_lists: list[list[tuple[int, int]]] = []
        for t in meta.and_tokens:
            pos = get_positions(t.word, t.is_full)
            if not pos:
                return False
            position_lists.append(pos)
        if not and_non_overlapping(position_lists):
            return False

    if meta.or_tokens:
        ok_or = False
        for t in meta.or_tokens:
            if get_positions(t.word, t.is_full):
                ok_or = True
                break
        if not ok_or:
            return False

    if meta.and_non_tokens:
        non_and_positions: list[list[tuple[int, int]]] = []
        for token in meta.and_non_tokens:
            pos = (
                get_positions(token, True)
                if compact_text == text
                else all_positions_full(compact_text, token)
            )
            if not pos:
                non_and_positions = []
                break
            non_and_positions.append(pos)
        if non_and_positions and and_non_overlapping(non_and_positions):
            return False

    if meta.or_non_tokens:
        for token in meta.or_non_tokens:
            if get_positions(token, True) if compact_text == text else all_positions_full(compact_text, token):
                return False

    return True


def _match_holding_to_base(
    holding_id: str,
    text_value: Any,
    candidate_indices: list[int] | None,
) -> SingleHoldingMatchResult:
    """Общая логика сопоставления холдинга: fix-ID + поиск по ключам."""
    text = normalize_text(text_value)
    if not text:
        empty = MATCH_STATUS_TEXTS.none
        return SingleHoldingMatchResult(empty, empty, empty, 0, (), ())

    compact_text = normalize_non_text(text)
    pos_cache: dict[tuple[str, bool], list[tuple[int, int]]] = {}

    def get_positions(word: str, is_full: bool) -> list[tuple[int, int]]:
        key = (word, is_full)
        cached = pos_cache.get(key)
        if cached is not None:
            return cached
        out = all_positions_full(text, word) if is_full else positions_not(text, word)
        pos_cache[key] = out
        return out

    fixed_indices: list[int] = []
    key_indices: list[int] = []
    matches: list[str] = []
    seen_indices: set[int] = set()

    for idx in FIX_INDICES_BY_HOLDING_ID.get(holding_id, ()):
        if idx in seen_indices:
            continue
        meta = BASE_METAS[idx]
        if holding_id in meta.fix_ids:
            seen_indices.add(idx)
            fixed_indices.append(idx)
            if meta.gsz_value:
                matches.append(meta.gsz_value)

    indices = candidate_indices if candidate_indices is not None else list(range(len(BASE_METAS)))
    for idx in indices:
        if idx in seen_indices:
            continue
        meta = BASE_METAS[idx]
        if meta.fix_mode == "resolved":
            continue
        if meta.fix_mode in ("fallback", "none", "partial"):
            if _meta_matches_with_cache(text, compact_text, meta, get_positions):
                seen_indices.add(idx)
                key_indices.append(idx)
                if meta.gsz_value:
                    matches.append(meta.gsz_value)

    matched_indices = tuple(fixed_indices + key_indices)
    fixed_tuple = tuple(fixed_indices)
    primary, debug = format_match_columns(matches, MATCH_STATUS_TEXTS)
    status = compute_holding_status_lines(len(matches), len(fixed_indices), MATCH_STATUS_TEXTS)
    return SingleHoldingMatchResult(
        primary=primary,
        debug=debug,
        status=status,
        count=len(matches),
        matched_indices=matched_indices,
        fixed_indices=fixed_tuple,
    )


def match_single_holding(holding_id: Any, text_value: Any) -> SingleHoldingMatchResult:
    """Быстрый поиск: fix-ID + якорный предфильтр + полная проверка кандидатов."""
    holding_id_norm = normalize_holding_id(holding_id)
    text = normalize_text(text_value)
    if not text:
        empty = MATCH_STATUS_TEXTS.none
        return SingleHoldingMatchResult(empty, empty, empty, 0, (), ())
    return _match_holding_to_base(
        holding_id_norm,
        text_value,
        candidate_indices_for_text(text),
    )


def match_single_holding_brute(holding_id: Any, text_value: Any) -> SingleHoldingMatchResult:
    """Полный перебор справочника без якорного предфильтра (для регресс-тестов)."""
    holding_id_norm = normalize_holding_id(holding_id)
    text = normalize_text(text_value)
    if not text:
        empty = MATCH_STATUS_TEXTS.none
        return SingleHoldingMatchResult(empty, empty, empty, 0, (), ())
    return _match_holding_to_base(holding_id_norm, text_value, None)


def ensure_columns(rows: list[dict[str, Any]], cols: list[str], where: str) -> None:
    """Проверка наличия обязательных колонок перед сопоставлением."""
    if not rows:
        raise ValueError(f"{where} пуста")
    available = set(rows[0].keys())
    missing = [c for c in cols if c not in available]
    if missing:
        raise ValueError(f"В {where} отсутствуют колонки: {missing}")


# =============================================================================
# Конфигурация выходных колонок Excel (key → name / width / wrap)
# =============================================================================
def output_columns_by_key(columns: tuple[OutputColumnSpec, ...]) -> dict[str, OutputColumnSpec]:
    """Словарь key -> спецификация колонки."""
    return {column.key: column for column in columns}


def output_columns_by_role(columns: tuple[OutputColumnSpec, ...]) -> dict[str, OutputColumnSpec]:
    """Алиас для обратной совместимости."""
    return output_columns_by_key(columns)


def output_column_names(columns: tuple[OutputColumnSpec, ...]) -> list[str]:
    """Имена добавляемых колонок в порядке конфигурации."""
    return [column.name for column in columns]


def _legacy_width_overrides(format_cfg: dict[str, Any]) -> dict[str, float]:
    """Старые ключи ширины из output_format для обратной совместимости."""
    overrides: dict[str, float] = {}
    legacy_map = {
        "gsz_primary": "holding_gsz_min_width",
        "gsz_debug": "holding_debug_min_width",
        "found_holding": "base_found_holding_min_width",
        "found_holding_debug": "base_found_holding_debug_min_width",
    }
    for role, legacy_key in legacy_map.items():
        if legacy_key in format_cfg:
            overrides[role] = float(format_cfg[legacy_key])
    return overrides


def _legacy_wrap_overrides(format_cfg: dict[str, Any]) -> dict[str, bool]:
    """Старые флаги переноса строк из output_format."""
    overrides: dict[str, bool] = {}
    if "holding_debug_wrap" in format_cfg:
        overrides["gsz_debug"] = bool(format_cfg["holding_debug_wrap"])
    if "base_found_holding_debug_wrap" in format_cfg:
        overrides["found_holding_debug"] = bool(format_cfg["base_found_holding_debug_wrap"])
    return overrides


def _build_output_column_spec(
    default_spec: OutputColumnSpec,
    item: Any,
    default_width: float,
    legacy_widths: dict[str, float],
    legacy_wraps: dict[str, bool],
) -> OutputColumnSpec:
    """Сборка спецификации колонки из элемента конфигурации."""
    if item is None:
        return OutputColumnSpec(
            key=default_spec.key,
            name=default_spec.name,
            width=float(
                legacy_widths.get(
                    default_spec.key,
                    default_spec.width if default_spec.width > 0 else default_width,
                )
            ),
            wrap=legacy_wraps.get(default_spec.key, default_spec.wrap),
        )

    if isinstance(item, str):
        return OutputColumnSpec(
            key=default_spec.key,
            name=item,
            width=default_width,
            wrap=legacy_wraps.get(default_spec.key, default_spec.wrap),
        )

    if isinstance(item, dict):
        return OutputColumnSpec(
            key=default_spec.key,
            name=str(item.get("name", default_spec.name)),
            width=float(item.get("width", default_width)),
            wrap=bool(item.get("wrap", legacy_wraps.get(default_spec.key, default_spec.wrap))),
        )

    raise ValueError(
        f"Некорректная настройка columns['{default_spec.key}'] в output_format: ожидается объект"
    )


def _parse_columns_mapping(
    raw_columns: dict[str, Any],
    defaults: tuple[OutputColumnSpec, ...],
    default_width: float,
    legacy_widths: dict[str, float],
    legacy_wraps: dict[str, bool],
) -> tuple[OutputColumnSpec, ...]:
    """Разбор columns как словаря key -> {name, width, wrap}."""
    allowed_keys = {spec.key for spec in defaults}
    unknown_keys = sorted(set(raw_columns) - allowed_keys)
    if unknown_keys:
        raise ValueError(
            f"Неизвестные ключи columns в output_format: {unknown_keys}. "
            f"Допустимые ключи: {sorted(allowed_keys)}"
        )

    return tuple(
        _build_output_column_spec(
            default_spec=default_spec,
            item=raw_columns.get(default_spec.key),
            default_width=default_width,
            legacy_widths=legacy_widths,
            legacy_wraps=legacy_wraps,
        )
        for default_spec in defaults
    )


def _parse_columns_list(
    raw_columns: list[Any],
    defaults: tuple[OutputColumnSpec, ...],
    default_width: float,
    legacy_widths: dict[str, float],
    legacy_wraps: dict[str, bool],
) -> tuple[OutputColumnSpec, ...]:
    """Разбор columns как списка (устаревший формат по позиции или с полем key)."""
    if raw_columns and all(isinstance(item, dict) and item.get("key") for item in raw_columns):
        keyed: dict[str, Any] = {}
        for item in raw_columns:
            if not isinstance(item, dict):
                continue
            keyed[str(item["key"])] = item
        return _parse_columns_mapping(keyed, defaults, default_width, legacy_widths, legacy_wraps)

    return tuple(
        _build_output_column_spec(
            default_spec=default_spec,
            item=raw_columns[idx] if idx < len(raw_columns) else None,
            default_width=default_width,
            legacy_widths=legacy_widths,
            legacy_wraps=legacy_wraps,
        )
        for idx, default_spec in enumerate(defaults)
    )


def parse_sheet_output_columns(
    sheet_cfg: Any,
    defaults: tuple[OutputColumnSpec, ...],
    default_width: float,
    legacy_widths: dict[str, float] | None = None,
    legacy_wraps: dict[str, bool] | None = None,
) -> tuple[OutputColumnSpec, ...]:
    """Разбор columns листа из config.json."""
    legacy_widths = legacy_widths or {}
    legacy_wraps = legacy_wraps or {}

    raw_columns: Any = None
    if isinstance(sheet_cfg, dict):
        raw_columns = sheet_cfg.get("columns")
    elif isinstance(sheet_cfg, list):
        raw_columns = sheet_cfg

    if raw_columns is None or raw_columns == []:
        return tuple(
            _build_output_column_spec(
                default_spec=spec,
                item=None,
                default_width=default_width,
                legacy_widths=legacy_widths,
                legacy_wraps=legacy_wraps,
            )
            for spec in defaults
        )

    if isinstance(raw_columns, dict):
        return _parse_columns_mapping(
            raw_columns=raw_columns,
            defaults=defaults,
            default_width=default_width,
            legacy_widths=legacy_widths,
            legacy_wraps=legacy_wraps,
        )

    if isinstance(raw_columns, list):
        return _parse_columns_list(
            raw_columns=raw_columns,
            defaults=defaults,
            default_width=default_width,
            legacy_widths=legacy_widths,
            legacy_wraps=legacy_wraps,
        )

    raise ValueError("output_format.columns должен быть объектом или списком")


def resolve_output_format(format_cfg: dict[str, Any] | None) -> dict[str, Any]:
    """Нормализация output_format: имена/ширины добавляемых колонок по листам."""
    cfg = dict(format_cfg or {})
    min_width_all = float(cfg.get("min_width_all", DEFAULT_MIN_WIDTH_ALL))
    legacy_widths = _legacy_width_overrides(cfg)
    legacy_wraps = _legacy_wrap_overrides(cfg)

    holding_columns = parse_sheet_output_columns(
        sheet_cfg=cfg.get("holding_sheet"),
        defaults=DEFAULT_HOLDING_OUTPUT_COLUMNS,
        default_width=min_width_all,
        legacy_widths=legacy_widths,
        legacy_wraps=legacy_wraps,
    )
    base_columns = parse_sheet_output_columns(
        sheet_cfg=cfg.get("base_sheet"),
        defaults=DEFAULT_BASE_OUTPUT_COLUMNS,
        default_width=min_width_all,
        legacy_widths=legacy_widths,
        legacy_wraps=legacy_wraps,
    )

    cfg["min_width_all"] = min_width_all
    cfg["holding_columns"] = holding_columns
    cfg["base_columns"] = base_columns
    return cfg


def reorder_row_with_output_columns(
    row: dict[str, Any],
    output_columns: tuple[OutputColumnSpec, ...],
) -> dict[str, Any]:
    """Исходные колонки + добавляемые в порядке из конфигурации."""
    output_names = set(output_column_names(output_columns))
    original_keys = [key for key in row.keys() if key not in output_names]
    ordered_keys = original_keys + [column.name for column in output_columns if column.name in row]
    return {key: row[key] for key in ordered_keys}


def apply_output_column_specs(
    ws: Any,
    output_columns: tuple[OutputColumnSpec, ...],
    min_width_all: float,
) -> None:
    """Минимальная ширина для всех колонок и точечные настройки добавляемых."""
    apply_min_column_widths(ws, min_width_all)
    for column in output_columns:
        apply_column_min_width_by_header(ws, column.name, column.width)
        if column.wrap:
            apply_wrap_for_column_by_header(ws, column.name)


# =============================================================================
# Запись листов Excel: данные + форматирование
# =============================================================================
def write_sheet(ws: Any, rows: list[dict[str, Any]]) -> None:
    """Запись list[dict] на лист: первая строка — заголовки, далее данные."""
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def apply_sheet_formatting(
    ws: Any,
    header_center: bool,
    header_wrap: bool,
    header_bold: bool,
    freeze_rows: int,
    freeze_cols: int,
    format_data_vertical_center: bool = True,
) -> None:
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.alignment = Alignment(
                horizontal="center" if header_center else cell.alignment.horizontal,
                vertical="center",
                wrap_text=header_wrap,
            )
            if header_bold:
                cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True)

    if format_data_vertical_center and ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                existing = cell.alignment if cell.alignment is not None else Alignment()
                cell.alignment = Alignment(
                    horizontal=existing.horizontal,
                    vertical="center",
                    text_rotation=existing.text_rotation,
                    wrap_text=existing.wrap_text,
                    shrink_to_fit=existing.shrink_to_fit,
                    indent=existing.indent,
                    relativeIndent=existing.relativeIndent,
                    justifyLastLine=existing.justifyLastLine,
                    readingOrder=existing.readingOrder,
                )

    # Автофильтр по всей области данных листа.
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    if freeze_rows > 0 or freeze_cols > 0:
        ws.freeze_panes = ws.cell(row=freeze_rows + 1, column=freeze_cols + 1)


def get_header_index_map(ws: Any) -> dict[str, int]:
    out: dict[str, int] = {}
    if ws.max_row < 1:
        return out
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val is not None:
            out[str(val)] = col
    return out


def apply_min_column_widths(ws: Any, min_width: float) -> None:
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        current = ws.column_dimensions[col_letter].width
        current_val = float(current) if current is not None else 0.0
        ws.column_dimensions[col_letter].width = max(min_width, current_val)


def apply_column_min_width_by_header(ws: Any, header: str, min_width: float) -> None:
    header_map = get_header_index_map(ws)
    col = header_map.get(header)
    if col is None:
        return
    col_letter = get_column_letter(col)
    current = ws.column_dimensions[col_letter].width
    current_val = float(current) if current is not None else 0.0
    ws.column_dimensions[col_letter].width = max(min_width, current_val)


def apply_wrap_for_column_by_header(ws: Any, header: str) -> None:
    header_map = get_header_index_map(ws)
    col = header_map.get(header)
    if col is None:
        return
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.alignment if cell.alignment is not None else Alignment()
        cell.alignment = Alignment(
            horizontal=existing.horizontal,
            vertical="center",
            text_rotation=existing.text_rotation,
            wrap_text=True,
            shrink_to_fit=existing.shrink_to_fit,
            indent=existing.indent,
            relativeIndent=existing.relativeIndent,
            justifyLastLine=existing.justifyLastLine,
            readingOrder=existing.readingOrder,
        )


# =============================================================================
# Обогащение листа _base_gsz: найденные холдинги и аналитика по ключам
# =============================================================================
def format_holding_entry(
    hold_row: dict[str, Any],
    holding_id_column: str,
    holding_name_column: str,
    with_trailing_semicolon: bool = False,
) -> str:
    """Форматирование холдинга как [ID]: наименование."""
    holding_id = hold_row.get(holding_id_column, "")
    holding_name = hold_row.get(holding_name_column, "")
    if holding_id is None:
        holding_id = ""
    if holding_name is None:
        holding_name = ""
    entry = f"[{holding_id}]: {holding_name}"
    if with_trailing_semicolon:
        entry += ";"
    return entry


def build_base_holding_match_columns(
    matched_pairs: list[tuple[int, bool]],
    hold_rows: list[dict[str, Any]],
    holding_id_column: str,
    holding_name_column: str,
    texts: MatchStatusTexts,
) -> tuple[str, str]:
    """Колонки «найденный холдинг» и отладка: fix-сначала, затем поиск по ключам."""
    fixed_indices = [hold_idx for hold_idx, is_fixed in matched_pairs if is_fixed]
    key_indices = [hold_idx for hold_idx, is_fixed in matched_pairs if not is_fixed]

    def entries_for(indices: list[int]) -> list[str]:
        result: list[str] = []
        for hold_idx in indices:
            if 0 <= hold_idx < len(hold_rows):
                result.append(
                    format_holding_entry(
                        hold_rows[hold_idx],
                        holding_id_column=holding_id_column,
                        holding_name_column=holding_name_column,
                        with_trailing_semicolon=False,
                    )
                )
        return result

    fixed_entries = entries_for(fixed_indices)
    key_entries = entries_for(key_indices)
    primary = compute_found_holding_primary(fixed_entries, key_entries, texts)

    debug_lines: list[str] = []
    for hold_idx in fixed_indices + key_indices:
        if 0 <= hold_idx < len(hold_rows):
            debug_lines.append(
                format_holding_entry(
                    hold_rows[hold_idx],
                    holding_id_column=holding_id_column,
                    holding_name_column=holding_name_column,
                    with_trailing_semicolon=True,
                )
            )
    debug_text = "\n".join(debug_lines) if debug_lines else texts.none
    return primary, debug_text


def reorder_base_row_columns(
    row: dict[str, Any],
    base_columns: tuple[OutputColumnSpec, ...],
) -> dict[str, Any]:
    """Служебные колонки _base_gsz — сразу после исходных, в порядке конфигурации."""
    return reorder_row_with_output_columns(row, base_columns)


def enrich_base_rows(
    base_rows: list[dict[str, Any]],
    base_metas: list[BaseMeta],
    all_key_cols: list[str],
    per_row_holding_counts: list[int],
    per_row_matched_holding_indices: list[list[tuple[int, bool]]],
    hold_rows: list[dict[str, Any]],
    holding_id_column: str,
    holding_name_column: str,
    base_columns: tuple[OutputColumnSpec, ...],
    status_texts: MatchStatusTexts | None = None,
) -> None:
    """Обратная проекция: для каждой строки _base_gsz заполняет служебные колонки.

    - holding_count — сколько холдингов сматчилось на эту строку;
    - found_holding / found_holding_debug — список холдингов в формате [ID]: имя;
    - match_status — текстовый статус сопоставления;
    - key_string / key_length / key_repeat_count — аналитика по конкатенации ключей.
    """
    texts = status_texts or MATCH_STATUS_TEXTS
    base_by_key = output_columns_by_key(base_columns)
    col_holding_count = base_by_key["holding_count"].name
    col_found_holding = base_by_key["found_holding"].name
    col_found_debug = base_by_key["found_holding_debug"].name
    col_status = base_by_key["match_status"].name
    col_key_string = base_by_key["key_string"].name
    col_key_length = base_by_key["key_length"].name
    col_key_repeat = base_by_key["key_repeat_count"].name

    # Первый проход: строка ключа и обратная проекция холдингов
    key_strings: list[str] = []
    for idx, row in enumerate(base_rows):
        parts: list[str] = []
        for col in all_key_cols:
            value = row.get(col)
            text = normalize_text(value)
            if text:
                parts.append(text)
        key_str = "_".join(parts)
        key_strings.append(key_str)

        matched_pairs = (
            per_row_matched_holding_indices[idx]
            if idx < len(per_row_matched_holding_indices)
            else []
        )
        found_primary, found_debug = build_base_holding_match_columns(
            matched_pairs=matched_pairs,
            hold_rows=hold_rows,
            holding_id_column=holding_id_column,
            holding_name_column=holding_name_column,
            texts=texts,
        )
        meta = base_metas[idx] if idx < len(base_metas) else BaseMeta("", False, (), (), (), (), None)
        row_status = compute_base_row_status_lines(meta, matched_pairs, texts)

        row[col_holding_count] = per_row_holding_counts[idx] if idx < len(per_row_holding_counts) else 0
        row[col_found_holding] = found_primary
        row[col_found_debug] = found_debug
        row[col_status] = row_status
        row[col_key_string] = key_str
        row[col_key_length] = len(key_str)

    # Второй проход: число повторов одинаковой строки ключа по всему справочнику
    freq: dict[str, int] = {}
    for ks in key_strings:
        if ks:
            freq[ks] = freq.get(ks, 0) + 1
    for idx, row in enumerate(base_rows):
        ks = key_strings[idx]
        row[col_key_repeat] = freq.get(ks, 0) if ks else 0
        base_rows[idx] = reorder_base_row_columns(row, base_columns)


def write_output_xlsx(
    output_path: Path,
    holding_rows: list[dict[str, Any]],
    base_rows: list[dict[str, Any]],
    holding_sheet: str,
    base_sheet: str,
    format_cfg: dict[str, Any] | None = None,
) -> None:
    """Создаёт книгу с двумя листами (холдинги + справочник) и применяет output_format."""
    from openpyxl import Workbook

    wb = Workbook()
    # Лист 1: холдинги с колонками условное ГСЗ / отладка / кол-во
    ws1 = wb.active
    ws1.title = holding_sheet[:31] if holding_sheet else "HOLD_OD"
    write_sheet(ws1, holding_rows)
    format_cfg = resolve_output_format(format_cfg)
    holding_columns = format_cfg["holding_columns"]
    base_columns = format_cfg["base_columns"]
    min_width_all = float(format_cfg["min_width_all"])

    apply_sheet_formatting(
        ws=ws1,
        header_center=bool(format_cfg.get("header_center", True)),
        header_wrap=bool(format_cfg.get("header_wrap", True)),
        header_bold=bool(format_cfg.get("header_bold", True)),
        freeze_rows=max(0, int(format_cfg.get("holding_freeze_rows", format_cfg.get("freeze_rows", 1)))),
        freeze_cols=max(0, int(format_cfg.get("holding_freeze_cols", format_cfg.get("freeze_cols", 3)))),
        format_data_vertical_center=bool(format_cfg.get("format_data_vertical_center", True)),
    )
    apply_output_column_specs(ws1, holding_columns, min_width_all)

    # Лист 2: справочник _base_gsz с обратной проекцией найденных холдингов
    ws2 = wb.create_sheet(title=base_sheet[:31] if base_sheet else "base_gsz")
    write_sheet(ws2, base_rows)
    apply_sheet_formatting(
        ws=ws2,
        header_center=bool(format_cfg.get("header_center", True)),
        header_wrap=bool(format_cfg.get("header_wrap", True)),
        header_bold=bool(format_cfg.get("header_bold", True)),
        freeze_rows=max(0, int(format_cfg.get("base_freeze_rows", 1))),
        freeze_cols=max(0, int(format_cfg.get("base_freeze_cols", 6))),
        format_data_vertical_center=bool(format_cfg.get("format_data_vertical_center", True)),
    )
    apply_output_column_specs(ws2, base_columns, min_width_all)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# =============================================================================
# Параллельная обработка батчей холдингов
# =============================================================================
def chunked(seq: list[Any], size: int) -> list[list[Any]]:
    if size <= 0:
        size = 1
    return [seq[i : i + size] for i in range(0, len(seq), size)]


def match_holding_batch(
    indexed_holding_batch: list[tuple[int, Any, Any]],
) -> tuple[list[tuple[str, str, str, int]], dict[int, list[tuple[int, bool]]]]:
    """Обработка батча холдингов в одном воркере.

    Возвращает результаты по холдингам и обратную проекцию:
    base_idx → список (индекс холдинга, совпадение по fix-ID).
    """
    rows_out: list[tuple[str, str, str, int]] = []
    row_holding_indices: dict[int, list[tuple[int, bool]]] = {}
    for hold_idx, holding_id, value in indexed_holding_batch:
        result = match_single_holding(holding_id, value)
        rows_out.append((result.primary, result.debug, result.status, result.count))
        fixed_set = set(result.fixed_indices)
        for base_idx in result.matched_indices:
            is_fixed = base_idx in fixed_set
            row_holding_indices.setdefault(base_idx, []).append((hold_idx, is_fixed))
    return rows_out, row_holding_indices


def short_text(value: Any, max_len: int = 80) -> str:
    s = str(value) if value is not None else ""
    s = s.replace("\n", " ").replace("\r", " ").strip()
    if len(s) <= max_len:
        return s
    return s[: max_len - 3] + "..."


def log(message: str) -> None:
    print(message, flush=True)
    if LOG_FILE_PATH is not None:
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with LOG_FILE_PATH.open("a", encoding="utf-8") as f:
            f.write(f"{ts} {message}\n")


def configure_unbuffered_console_output() -> None:
    # Для IDE/раннеров, где stdout может буферизоваться даже с flush.
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(line_buffering=True, write_through=True)
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(line_buffering=True, write_through=True)


def configure_file_logging(logs_dir: str, log_file_prefix: str) -> Path:
    global LOG_FILE_PATH

    root = Path(__file__).resolve().parent.parent
    target_dir = (root / logs_dir).resolve() if not Path(logs_dir).is_absolute() else Path(logs_dir).resolve()
    target_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_prefix = re.sub(r"[^a-zA-Z0-9_.-]+", "_", log_file_prefix).strip("_") or "gsz_matcher_parallel"
    LOG_FILE_PATH = target_dir / f"{safe_prefix}_{ts}.log"
    return LOG_FILE_PATH


def build_progress_message(
    processed: int,
    total_holdings: int,
    batch_idx: int,
    total_batches: int,
    match_started: float,
    holding_texts: list[Any],
    show_current_holding: bool,
    prefix: str = "[progress]",
) -> str:
    elapsed = time.perf_counter() - match_started
    speed = processed / elapsed if elapsed > 0 else 0.0
    remaining = max(0, total_holdings - processed)
    eta_sec = (remaining / speed) if speed > 0 else 0.0
    pct = (processed / total_holdings * 100.0) if total_holdings else 100.0
    msg = (
        f"{prefix} {processed}/{total_holdings} ({pct:.1f}%), "
        f"batch={batch_idx}/{total_batches}, speed={speed:.1f} hold/s, "
        f"eta={eta_sec:.1f}s"
    )
    if show_current_holding and processed > 0:
        msg += f", current='{short_text(holding_texts[processed - 1])}'"
    return msg


def build_batch_message(
    completed_batches: int,
    total_batches: int,
    work_batch_size: int,
    total_holdings: int,
    pending_futures: int,
) -> str:
    approx_done = min(total_holdings, completed_batches * work_batch_size)
    pct = (approx_done / total_holdings * 100.0) if total_holdings else 100.0
    return (
        f"[progress-batch] batches={completed_batches}/{total_batches}, "
        f"approx_holdings_done~={approx_done}/{total_holdings} ({pct:.1f}%), "
        f"in_flight={pending_futures}"
    )


def make_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Параллельное сопоставление холдингов и условных ГСЗ")
    p.add_argument("--input-xlsx", help="Путь к исходной Excel-книге")
    p.add_argument("--output-xlsx", help="Путь к выходному Excel-файлу")
    p.add_argument("--holding-table", help="Имя таблицы холдингов")
    p.add_argument("--base-table", help="Имя таблицы справочника ГСЗ")
    p.add_argument("--holding-column", help="Колонка с текстом холдинга")
    p.add_argument("--holding-id-column", help="Колонка с ID холдинга")
    p.add_argument("--gsz-column", help="Колонка значения условного ГСЗ")
    p.add_argument("--and-full-cols", help="AND full колонки через запятую")
    p.add_argument("--and-not-cols", help="AND not колонки через запятую")
    p.add_argument("--and-non-cols", help="AND non колонки через запятую")
    p.add_argument("--or-full-cols", help="OR full колонки через запятую")
    p.add_argument("--or-not-cols", help="OR not колонки через запятую")
    p.add_argument("--or-non-cols", help="OR non колонки через запятую")
    p.add_argument("--workers", type=int, help="Число процессов")
    p.add_argument("--work-batch-size", type=int, help="Размер батча холдингов на один worker")
    p.add_argument("--chunk-size", type=int, help=argparse.SUPPRESS)
    p.add_argument("--config-json", default=str(DEFAULT_CONFIG_PATH), help="JSON-файл с параметрами")
    return p


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"config-json не найден: {path}")
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def resolve_settings(args: argparse.Namespace) -> dict[str, Any]:
    config_path = Path(args.config_json).expanduser().resolve()
    cfg = load_config(config_path)
    block = cfg.get("gsz_matcher_parallel", {})

    defaults: dict[str, Any] = {
        "input_xlsx": block.get("input_xlsx"),
        "output_xlsx": block.get("output_xlsx"),
        "holding_table": block.get("holding_table", "_HOLD_OD"),
        "base_table": block.get("base_table", "_base_gsz"),
        "holding_column": block.get("holding_column", "Холдинг"),
        "holding_id_column": block.get("holding_id_column", DEFAULT_HOLDING_ID_COLUMN),
        "gsz_column": block.get("gsz_column", "Наименование, регион"),
        "and_full_cols": block.get("and_full_cols", DEFAULT_AND_FULL),
        "and_not_cols": block.get("and_not_cols", DEFAULT_AND_NOT),
        "and_non_cols": block.get("and_non_cols", DEFAULT_AND_NON),
        "or_full_cols": block.get("or_full_cols", DEFAULT_OR_FULL),
        "or_not_cols": block.get("or_not_cols", DEFAULT_OR_NOT),
        "or_non_cols": block.get("or_non_cols", DEFAULT_OR_NON),
        "fix_id_col": block.get("fix_id_col", DEFAULT_FIX_ID_COL),
        "match_status_texts": load_match_status_texts(block),
        "workers": block.get("workers", max(1, (mp.cpu_count() or 2) - 1)),
        "work_batch_size": block.get("work_batch_size", 50),
        "log_stages": block.get("log_stages", True),
        "progress_every_holdings": block.get("progress_every_holdings", 1000),
        "progress_every_base_rows": block.get("progress_every_base_rows", 1000),
        "progress_every_batches": block.get("progress_every_batches", 25),
        "progress_every_read_rows": block.get("progress_every_read_rows", 1000),
        "heartbeat_seconds": block.get("heartbeat_seconds", 10),
        "show_current_holding": block.get("show_current_holding", True),
        "diagnose_workbook_objects": block.get("diagnose_workbook_objects", True),
        "log_to_file": block.get("log_to_file", True),
        "logs_dir": block.get("logs_dir", "LOGS"),
        "log_file_prefix": block.get("log_file_prefix", "gsz_matcher_parallel"),
        "output_format": resolve_output_format(
            block.get(
                "output_format",
                {
                    "header_center": True,
                    "header_wrap": True,
                    "header_bold": True,
                    "holding_freeze_rows": 1,
                    "holding_freeze_cols": 3,
                    "base_freeze_rows": 1,
                    "base_freeze_cols": 3,
                    "min_width_all": DEFAULT_MIN_WIDTH_ALL,
                    "format_data_vertical_center": True,
                },
            )
        ),
        "output_add_timestamp": block.get("output_add_timestamp", True),
        "output_timestamp_format": block.get("output_timestamp_format", "%Y%m%d_%H%M%S"),
        "_config_dir": str(config_path.parent),
    }

    # Обратная совместимость: можно использовать старые плоские ключи.
    if defaults["input_xlsx"] is None and cfg.get("input_xlsx"):
        defaults["input_xlsx"] = cfg.get("input_xlsx")
    if defaults["output_xlsx"] is None and cfg.get("output_xlsx"):
        defaults["output_xlsx"] = cfg.get("output_xlsx")

    cli_overrides = {
        "input_xlsx": args.input_xlsx,
        "output_xlsx": args.output_xlsx,
        "holding_table": args.holding_table,
        "base_table": args.base_table,
        "holding_column": args.holding_column,
        "holding_id_column": args.holding_id_column,
        "gsz_column": args.gsz_column,
        "and_full_cols": args.and_full_cols,
        "and_not_cols": args.and_not_cols,
        "and_non_cols": args.and_non_cols,
        "or_full_cols": args.or_full_cols,
        "or_not_cols": args.or_not_cols,
        "or_non_cols": args.or_non_cols,
        "workers": args.workers,
        "work_batch_size": args.work_batch_size if args.work_batch_size is not None else args.chunk_size,
    }
    for key, value in cli_overrides.items():
        if value is not None:
            defaults[key] = value

    if not defaults["input_xlsx"] or not defaults["output_xlsx"]:
        raise ValueError(
            "Не заданы input/output пути для Python-матчера. "
            "Укажите их в config.json (блок gsz_matcher_parallel) "
            "или передайте --input-xlsx и --output-xlsx."
        )
    return defaults


def resolve_path(value: str, base_dir: Path) -> Path:
    p = Path(value).expanduser()
    if p.is_absolute():
        return p.resolve()
    return (base_dir / p).resolve()


def with_timestamp_suffix(path: Path, pattern: str = "%Y%m%d_%H%M%S") -> Path:
    ts = datetime.now().strftime(pattern)
    if path.suffix:
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")
    return path.with_name(f"{path.name}_{ts}")


# =============================================================================
# main: чтение → сопоставление → обогащение → запись Excel
# =============================================================================
def main() -> None:
    """Точка входа CLI: config → чтение Excel → сопоставление → запись результата."""
    configure_unbuffered_console_output()
    parser = make_arg_parser()
    args = parser.parse_args()
    settings = resolve_settings(args)
    config_dir = Path(str(settings["_config_dir"])).resolve()
    global MATCH_STATUS_TEXTS
    status_texts: MatchStatusTexts = settings["match_status_texts"]
    MATCH_STATUS_TEXTS = status_texts

    if settings["log_to_file"]:
        logs_dir_value = str(settings["logs_dir"])
        logs_dir_path = resolve_path(logs_dir_value, config_dir)
        log_path = configure_file_logging(
            logs_dir=str(logs_dir_path),
            log_file_prefix=str(settings["log_file_prefix"]),
        )
        log(f"[stage] File log: {log_path}")

    input_xlsx = resolve_path(str(settings["input_xlsx"]), config_dir)
    output_xlsx = resolve_path(str(settings["output_xlsx"]), config_dir)
    if settings["output_add_timestamp"]:
        output_xlsx = with_timestamp_suffix(
            output_xlsx,
            pattern=str(settings["output_timestamp_format"]),
        )

    and_full_cols = parse_cols(settings["and_full_cols"])
    and_not_cols = parse_cols(settings["and_not_cols"])
    and_non_cols = parse_cols(settings["and_non_cols"])
    or_full_cols = parse_cols(settings["or_full_cols"])
    or_not_cols = parse_cols(settings["or_not_cols"])
    or_non_cols = parse_cols(settings["or_non_cols"])

    t0 = time.perf_counter()
    if settings["log_stages"]:
        log(f"[stage] Script={Path(__file__).resolve()} pid={os.getpid()}")
        log("[stage] Запуск Python-матчера.")
        log(
            f"[stage] Конфиг: workers={settings['workers']}, "
            f"work_batch_size={settings['work_batch_size']}, "
            f"progress_every={settings['progress_every_holdings']}, "
            f"heartbeat={settings['heartbeat_seconds']}s"
        )
        if settings["diagnose_workbook_objects"]:
            log(f"[stage] Диагностика объектов книги: {input_xlsx}")
            info = inspect_workbook_objects(input_xlsx)
            log(f"[diag] Smart Tables ({len(info['table_names'])}): {', '.join(info['table_names']) or '-'}")
            log(f"[diag] Defined Names ({len(info['defined_names'])}): {', '.join(info['defined_names']) or '-'}")
            hold_as_table = settings["holding_table"] in set(info["table_names"])
            base_as_table = settings["base_table"] in set(info["table_names"])
            hold_as_name = settings["holding_table"] in set(info["defined_names"])
            base_as_name = settings["base_table"] in set(info["defined_names"])
            log(
                f"[diag] holding_table='{settings['holding_table']}': "
                f"smart_table={hold_as_table}, defined_name={hold_as_name}"
            )
            log(
                f"[diag] base_table='{settings['base_table']}': "
                f"smart_table={base_as_table}, defined_name={base_as_name}"
            )
            if (hold_as_name and not hold_as_table) or (base_as_name and not base_as_table):
                log(
                    "[diag-warning] Найдено имя как Defined Name, но не как Smart Table. "
                    "Скрипт читает только Smart Table (ListObject)."
                )
    # --- Этап 1: чтение смарт-таблиц из входной книги ---
    if settings["log_stages"]:
        log(f"[stage] Чтение таблицы {settings['holding_table']}...")
    hold_rows = read_excel_table(
        input_xlsx,
        settings["holding_table"],
        log_enabled=bool(settings["log_stages"]),
        progress_every_read_rows=max(1, int(settings["progress_every_read_rows"])),
    )
    if settings["log_stages"]:
        log(f"[stage] Таблица {settings['holding_table']} загружена: {len(hold_rows)} строк.")
        log(f"[stage] Чтение таблицы {settings['base_table']}...")
    base_rows = read_excel_table(
        input_xlsx,
        settings["base_table"],
        log_enabled=bool(settings["log_stages"]),
        progress_every_read_rows=max(1, int(settings["progress_every_read_rows"])),
    )
    if settings["log_stages"]:
        log(f"[stage] Таблица {settings['base_table']} загружена: {len(base_rows)} строк.")

    if settings["log_stages"]:
        log("[stage] Проверка обязательных колонок...")
    ensure_columns(
        hold_rows,
        [settings["holding_column"], settings["holding_id_column"]],
        f"таблице {settings['holding_table']}",
    )
    ensure_columns(
        base_rows,
        [settings["gsz_column"]]
        + and_full_cols
        + and_not_cols
        + and_non_cols
        + or_full_cols
        + or_not_cols
        + or_non_cols,
        f"таблице {settings['base_table']}",
    )

    # --- Этап 2: предобработка справочника (токены, якорь, non, fix-ID) один раз ---
    holdings_id_set = frozenset(
        normalize_holding_id(r.get(settings["holding_id_column"]))
        for r in hold_rows
        if normalize_holding_id(r.get(settings["holding_id_column"]))
    )
    fix_id_col = str(settings.get("fix_id_col", DEFAULT_FIX_ID_COL))
    if settings["log_stages"]:
        log("[stage] Подготовка метаданных справочника...")
    metas_list: list[BaseMeta] = []
    base_progress_every = max(1, int(settings["progress_every_base_rows"]))
    for idx, r in enumerate(base_rows, start=1):
        metas_list.append(
            build_meta_row(
                row=r,
                gsz_col=settings["gsz_column"],
                and_full_cols=and_full_cols,
                and_not_cols=and_not_cols,
                and_non_cols=and_non_cols,
                or_full_cols=or_full_cols,
                or_not_cols=or_not_cols,
                or_non_cols=or_non_cols,
                fix_id_col=fix_id_col,
                holdings_id_set=holdings_id_set,
            )
        )
        if settings["log_stages"] and (idx % base_progress_every == 0 or idx == len(base_rows)):
            log(f"[progress-base] {idx}/{len(base_rows)}")
    metas = tuple(metas_list)
    base_holding_counts = [0] * len(base_rows)
    base_matched_holding_indices: list[list[tuple[int, bool]]] = [[] for _ in base_rows]

    total_holdings = len(hold_rows)
    total_base = len(base_rows)
    approx_comparisons = total_holdings * total_base
    if settings["log_stages"]:
        log("[stage] Подготовка справочника завершена.")
        log(
            f"[stage] Оценка масштаба: {total_holdings} холдингов x "
            f"{total_base} строк справочника ~= {approx_comparisons} проверок"
        )

    holding_texts = [r.get(settings["holding_column"]) for r in hold_rows]
    holding_ids = [
        normalize_holding_id(r.get(settings["holding_id_column"])) for r in hold_rows
    ]
    work_batch_size = max(1, int(settings["work_batch_size"]))
    progress_every = max(1, int(settings["progress_every_holdings"]))
    progress_every_batches = max(1, int(settings["progress_every_batches"]))
    workers = max(1, int(settings["workers"]))
    heartbeat_seconds = max(1, int(settings["heartbeat_seconds"]))
    indexed_holding_batch = [
        (idx, holding_ids[idx], holding_texts[idx]) for idx in range(len(hold_rows))
    ]
    batches = chunked(indexed_holding_batch, work_batch_size)

    # --- Этап 3: параллельное сопоставление холдингов батчами ---
    if settings["log_stages"]:
        log(
            f"[stage] Старт сопоставления: workers={workers}, work_batch={work_batch_size}, "
            f"progress_every={progress_every}, "
            f"progress_every_batches={progress_every_batches}"
        )

    results: list[tuple[str, str, str, int]] = []
    processed = 0
    next_progress = progress_every
    match_started = time.perf_counter()
    with ProcessPoolExecutor(
        max_workers=workers,
        initializer=worker_init,
        initargs=(metas, status_texts),
    ) as ex:
        future_to_idx = {ex.submit(match_holding_batch, b): i for i, b in enumerate(batches)}
        ordered_batches: dict[
            int,
            tuple[list[tuple[str, str, str, int]], dict[int, list[tuple[int, bool]]]],
        ] = {}
        next_idx_to_flush = 0
        completed_batches = 0
        next_batch_progress = progress_every_batches
        pending = set(future_to_idx.keys())
        while pending:
            done, pending = wait(pending, timeout=heartbeat_seconds, return_when=FIRST_COMPLETED)
            if not done:
                log(
                    build_progress_message(
                        processed=processed,
                        total_holdings=total_holdings,
                        batch_idx=next_idx_to_flush,
                        total_batches=len(batches),
                        match_started=match_started,
                        holding_texts=holding_texts,
                        show_current_holding=settings["show_current_holding"],
                        prefix="[heartbeat]",
                    )
                )
                continue

            for fut in done:
                idx = future_to_idx[fut]
                ordered_batches[idx] = fut.result()
                completed_batches += 1
                if completed_batches >= next_batch_progress or completed_batches == len(batches):
                    log(
                        build_batch_message(
                            completed_batches=completed_batches,
                            total_batches=len(batches),
                            work_batch_size=work_batch_size,
                            total_holdings=total_holdings,
                            pending_futures=len(pending),
                        )
                    )
                    while next_batch_progress <= completed_batches:
                        next_batch_progress += progress_every_batches

            while next_idx_to_flush in ordered_batches:
                batch_result, batch_holding_indices = ordered_batches.pop(next_idx_to_flush)
                results.extend(batch_result)
                processed += len(batch_result)
                for row_idx, hold_pairs in batch_holding_indices.items():
                    if 0 <= row_idx < len(base_holding_counts):
                        base_holding_counts[row_idx] += len(hold_pairs)
                        base_matched_holding_indices[row_idx].extend(hold_pairs)

                if processed >= next_progress or processed == total_holdings:
                    log(
                        build_progress_message(
                            processed=processed,
                            total_holdings=total_holdings,
                            batch_idx=next_idx_to_flush + 1,
                            total_batches=len(batches),
                            match_started=match_started,
                            holding_texts=holding_texts,
                            show_current_holding=settings["show_current_holding"],
                            prefix="[progress]",
                        )
                    )
                    while next_progress <= processed:
                        next_progress += progress_every
                next_idx_to_flush += 1

    # --- Этап 4: заполнение выходных колонок на обоих листах ---
    output_format = settings["output_format"]
    holding_by_key = output_columns_by_key(output_format["holding_columns"])
    for row, res in zip(hold_rows, results):
        row[holding_by_key["gsz_primary"].name] = res[0]
        row[holding_by_key["gsz_debug"].name] = res[1]
        row[holding_by_key["match_status"].name] = res[2]
        row[holding_by_key["match_count"].name] = res[3]

    for idx, row in enumerate(hold_rows):
        hold_rows[idx] = reorder_row_with_output_columns(row, output_format["holding_columns"])

    all_key_cols = and_full_cols + and_not_cols + and_non_cols + or_full_cols + or_not_cols + or_non_cols
    enrich_base_rows(
        base_rows,
        base_metas=metas_list,
        all_key_cols=all_key_cols,
        per_row_holding_counts=base_holding_counts,
        per_row_matched_holding_indices=base_matched_holding_indices,
        hold_rows=hold_rows,
        holding_id_column=str(settings["holding_id_column"]),
        holding_name_column=str(settings["holding_column"]),
        base_columns=output_format["base_columns"],
        status_texts=status_texts,
    )

    # --- Этап 5: запись Excel с форматированием из output_format ---
    if settings["log_stages"]:
        log("[stage] Запись результата в Excel...")
    write_output_xlsx(
        output_path=output_xlsx,
        holding_rows=hold_rows,
        base_rows=base_rows,
        holding_sheet=settings["holding_table"],
        base_sheet=settings["base_table"],
        format_cfg=output_format,
    )

    t1 = time.perf_counter()
    if settings["log_stages"]:
        log("[stage] Готово.")
    log(f"Готово. Вход: {input_xlsx}")
    log(f"Результат: {output_xlsx}")
    log(f"Холдингов: {len(hold_rows)}, строк _base_gsz: {len(base_rows)}")
    log(
        f"Потоков: {max(1, int(settings['workers']))}, "
        f"work_batch: {max(1, int(settings['work_batch_size']))}"
    )
    log(f"Время: {t1 - t0:.2f} сек")


if __name__ == "__main__":
    main()
