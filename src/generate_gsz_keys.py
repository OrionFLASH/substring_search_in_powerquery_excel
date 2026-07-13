#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Генерация ключей для таблицы _base_gsz из Excel-книги.

Источник: смарт-таблица `_base_gsz` в файле `gsz_matcher_parallel.input_xlsx`,
колонка `gsz_column` (обычно «Наименование, регион»).

Правила ключей:
- только слова/буквы из исходного короткого имени (до запятой);
- допускается транслитерация RU↔EN того же токена;
- выдуманные слова и символы запрещены;
- отдельные буквы (например «А» в «Строй-А») → `or_not` (+ транслит «A»).

Состав колонок: `and_full_cols`, `and_not_cols`, `and_non_cols`,
`or_full_cols`, `or_not_cols`, `or_non_cols` из config.json.
"""

from __future__ import annotations

import json
import logging
import re
import sys
from datetime import datetime
from dataclasses import dataclass, field
from itertools import combinations
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

КОРЕНЬ = Path(__file__).resolve().parent.parent
КОНФИГ_ПУТЬ = КОРЕНЬ / "config.json"
SRC = КОРЕНЬ / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from gsz_matcher_parallel import (  # noqa: E402
    DEFAULT_AND_FULL,
    DEFAULT_AND_NON,
    DEFAULT_AND_NOT,
    DEFAULT_OR_FULL,
    DEFAULT_OR_NON,
    DEFAULT_OR_NOT,
    read_excel_table,
)

# Юридические/служебные формы — не ключи (кроме резервного режима)
СТОП_СЛОВА = {
    "гк",
    "группа",
    "группы",
    "холдинг",
    "ооо",
    "ао",
    "зао",
    "пао",
    "ск",
    "сз",
    "кп",
    "ак",
    "фонд",
    "компания",
    "корпорация",
    "концерн",
    "объединение",
    "застройщиков",
    "застройщик",
    "девелопмент",
    "development",
    "group",
    "holding",
    "строительный",
    "трест",
    "нпф",
    "мк",
    "нп",
    "ип",
}

ОБЩИЕ_ТОКЕНЫ = {
    "group",
    "development",
    "инвест",
    "дом",
    "сити",
    "плюс",
    "регион",
    "север",
    "юг",
    "восток",
    "запад",
    "центр",
    "новый",
    "новая",
    "проект",
}

_CYR_TO_LAT: dict[str, str] = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}
_LAT_TO_CYR: dict[str, str] = {
    "a": "а",
    "b": "б",
    "c": "с",
    "d": "д",
    "e": "е",
    "f": "ф",
    "g": "г",
    "h": "х",
    "i": "и",
    "j": "й",
    "k": "к",
    "l": "л",
    "m": "м",
    "n": "н",
    "o": "о",
    "p": "п",
    "q": "к",
    "r": "р",
    "s": "с",
    "t": "т",
    "u": "у",
    "v": "в",
    "w": "в",
    "x": "кс",
    "y": "й",
    "z": "з",
}


@dataclass
class КолонкиКлючей:
    and_full: list[str]
    and_not: list[str]
    and_non: list[str]
    or_full: list[str]
    or_not: list[str]
    or_non: list[str]

    def все(self) -> list[str]:
        return (
            list(self.and_full)
            + list(self.and_not)
            + list(self.and_non)
            + list(self.or_full)
            + list(self.or_not)
            + list(self.or_non)
        )


@dataclass
class КлючиСтроки:
    and_full: list[str] = field(default_factory=list)
    and_not: list[str] = field(default_factory=list)
    and_non: list[str] = field(default_factory=list)
    or_full: list[str] = field(default_factory=list)
    or_not: list[str] = field(default_factory=list)
    or_non: list[str] = field(default_factory=list)
    дубликат: bool = False
    комментарий: str = ""

    def в_словарь(self, колонки: КолонкиКлючей) -> dict[str, str]:
        результат: dict[str, str] = {к: "" for к in колонки.все()}

        def _заполнить(имена: list[str], значения: list[str]) -> None:
            for i, имя in enumerate(имена):
                if i < len(значения) and значения[i]:
                    результат[имя] = значения[i]

        _заполнить(колонки.and_full, self.and_full)
        _заполнить(колонки.and_not, self.and_not)
        _заполнить(колонки.and_non, self.and_non)
        _заполнить(колонки.or_full, self.or_full)
        _заполнить(колонки.or_not, self.or_not)
        _заполнить(колонки.or_non, self.or_non)
        return результат


@dataclass
class ЗаписьГСЗ:
    полное_имя: str
    короткое_имя: str
    токены: list[str]
    буквы: list[str] = field(default_factory=list)
    ключи: КлючиСтроки = field(default_factory=КлючиСтроки)


def настроить_логирование() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - [%(levelname)s] - %(message)s")


def resolve_path(value: str, base_dir: Path) -> Path:
    p = Path(value).expanduser()
    return p.resolve() if p.is_absolute() else (base_dir / p).resolve()


def загрузить_конфиг() -> dict[str, Any]:
    if not КОНФИГ_ПУТЬ.exists():
        raise FileNotFoundError(f"config.json не найден: {КОНФИГ_ПУТЬ}")
    with КОНФИГ_ПУТЬ.open(encoding="utf-8") as файл:
        данные = json.load(файл)
    блок = данные.get("gsz_matcher_parallel", {})
    колонки = КолонкиКлючей(
        and_full=list(блок.get("and_full_cols", DEFAULT_AND_FULL)),
        and_not=list(блок.get("and_not_cols", DEFAULT_AND_NOT)),
        and_non=list(блок.get("and_non_cols", DEFAULT_AND_NON)),
        or_full=list(блок.get("or_full_cols", DEFAULT_OR_FULL)),
        or_not=list(блок.get("or_not_cols", DEFAULT_OR_NOT)),
        or_non=list(блок.get("or_non_cols", DEFAULT_OR_NON)),
    )
    return {
        "input_xlsx": блок.get("input_xlsx", "input/Tables+Holding.xlsx"),
        "base_table": блок.get("base_table", "_base_gsz"),
        "gsz_column": блок.get("gsz_column", "Наименование, регион"),
        "output_file": данные.get("output_file", "output/_base_gsz_keys.xlsx"),
        "output_add_timestamp": bool(данные.get("output_add_timestamp", True)),
        "output_timestamp_format": str(
            данные.get("output_timestamp_format")
            or блок.get("output_timestamp_format")
            or "%Y%m%d_%H%M%S"
        ),
        "min_token_length": int(данные.get("min_token_length", 2)),
        "short_token_not_max_len": int(данные.get("short_token_not_max_len", 5)),
        "common_token_threshold": int(данные.get("common_token_threshold", 80)),
        "колонки": колонки,
        "_config_dir": str(КОНФИГ_ПУТЬ.parent),
    }


def with_timestamp_suffix(path: Path, pattern: str = "%Y%m%d_%H%M%S") -> Path:
    """Добавить таймштамп к имени файла перед расширением."""
    ts = datetime.now().strftime(pattern)
    if path.suffix:
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")
    return path.with_name(f"{path.name}_{ts}")


def нормализовать(текст: str) -> str:
    return текст.strip().lower()


def нормализовать_non(текст: str) -> str:
    return "".join(ch for ch in нормализовать(текст) if not ch.isspace())


def это_буква(символ: str) -> bool:
    if len(символ) != 1:
        return False
    с = символ.lower()
    return ("a" <= с <= "z") or ("а" <= с <= "я") or с == "ё"


def все_позиции_full(текст: str, слово: str) -> list[tuple[int, int]]:
    if not слово or not текст or len(слово) > len(текст):
        return []
    дл = len(слово)
    return [(i, i + дл - 1) for i in range(len(текст) - дл + 1) if текст[i : i + дл] == слово]


def позиции_not(текст: str, слово: str) -> list[tuple[int, int]]:
    результат: list[tuple[int, int]] = []
    for старт, конец in все_позиции_full(текст, слово):
        слева = старт == 0 or not это_буква(текст[старт - 1])
        справа = конец == len(текст) - 1 or not это_буква(текст[конец + 1])
        if слева and справа:
            результат.append((старт, конец))
    return результат


def интервалы_пересекаются(а: tuple[int, int], б: tuple[int, int]) -> bool:
    return not (а[1] < б[0] or б[1] < а[0])


def подобрать_без_пересечения(
    списки: list[list[tuple[int, int]]],
    индекс: int = 0,
    выбранные: list[tuple[int, int]] | None = None,
) -> bool:
    выбранные = выбранные or []
    if индекс >= len(списки):
        return True
    for интервал in списки[индекс]:
        if any(интервалы_пересекаются(интервал, в) for в in выбранные):
            continue
        if подобрать_без_пересечения(списки, индекс + 1, выбранные + [интервал]):
            return True
    return False


def проверить_and(текст: str, токены: list[tuple[str, bool]]) -> bool:
    if not токены:
        return True
    норм = нормализовать(текст)
    списки: list[list[tuple[int, int]]] = []
    for слово, is_full in токены:
        w = нормализовать(слово)
        if not w:
            return False
        позиции = все_позиции_full(норм, w) if is_full else позиции_not(норм, w)
        if not позиции:
            return False
        списки.append(позиции)
    return подобрать_без_пересечения(списки)


def проверить_or(текст: str, токены: list[tuple[str, bool]]) -> bool:
    if not токены:
        return True
    норм = нормализовать(текст)
    for слово, is_full in токены:
        w = нормализовать(слово)
        if not w:
            continue
        позиции = все_позиции_full(норм, w) if is_full else позиции_not(норм, w)
        if позиции:
            return True
    return False


def проверить_and_non(текст: str, токены: list[str]) -> bool:
    if not токены:
        return False
    compact = нормализовать_non(текст)
    списки: list[list[tuple[int, int]]] = []
    for token in токены:
        w = нормализовать_non(token)
        if not w:
            return False
        pos = все_позиции_full(compact, w)
        if not pos:
            return False
        списки.append(pos)
    return подобрать_без_пересечения(списки)


def проверить_or_non(текст: str, токены: list[str]) -> bool:
    if not токены:
        return False
    compact = нормализовать_non(текст)
    for token in токены:
        w = нормализовать_non(token)
        if w and все_позиции_full(compact, w):
            return True
    return False


def строка_совпала(текст_холдинга: str, ключи: КлючиСтроки) -> bool:
    if not нормализовать(текст_холдинга):
        return False
    and_токены = [(w, True) for w in ключи.and_full] + [(w, False) for w in ключи.and_not]
    or_токены = [(w, True) for w in ключи.or_full] + [(w, False) for w in ключи.or_not]
    if not (and_токены or or_токены or ключи.and_non or ключи.or_non):
        return False
    if and_токены and not проверить_and(текст_холдинга, and_токены):
        return False
    if or_токены and not проверить_or(текст_холдинга, or_токены):
        return False
    if not and_токены and not or_токены:
        return False
    if проверить_and_non(текст_холдинга, ключи.and_non):
        return False
    if проверить_or_non(текст_холдинга, ключи.or_non):
        return False
    return True


def разобрать_части_имени(короткое_имя: str) -> list[str]:
    текст = короткое_имя.strip()
    скобки = re.findall(r"\(([^)]+)\)", текст)
    текст = re.sub(r"\([^)]*\)", " ", текст)
    текст = re.sub(r"[®™]", "", текст)
    части = re.split(r"[\s\-–—/]+", текст)
    части.extend(re.split(r"[\s\-–—/]+", " ".join(скобки)))
    очищенные: list[str] = []
    for часть in части:
        очищенная = re.sub(r"[^\w]", "", часть, flags=re.UNICODE)
        if очищенная:
            очищенные.append(очищенная)
    return очищенные


def транслит_в_лат(текст: str) -> str:
    out: list[str] = []
    for ch in текст.lower():
        if ("а" <= ch <= "я") or ch == "ё":
            out.append(_CYR_TO_LAT.get(ch, ch))
        else:
            out.append(ch)
    return "".join(out)


def транслит_в_кир(текст: str) -> str:
    out: list[str] = []
    for ch in текст.lower():
        if "a" <= ch <= "z":
            out.append(_LAT_TO_CYR.get(ch, ch))
        else:
            out.append(ch)
    return "".join(out)


def варианты_токена(токен: str) -> list[str]:
    base = токен.strip()
    if not base:
        return []
    варианты = [base]
    lat = транслит_в_лат(base)
    cyr = транслит_в_кир(base)
    if lat and lat.lower() != base.lower():
        варианты.append(lat)
    if cyr and cyr.lower() != base.lower():
        варианты.append(cyr)
    уникальные: list[str] = []
    видели: set[str] = set()
    for в in варианты:
        кл = в.lower()
        if кл and кл not in видели:
            видели.add(кл)
            уникальные.append(в)
    return уникальные


def ключ_обоснован_именем(короткое_имя: str, ключ: str) -> bool:
    """Ключ только из части имени или транслита части имени."""
    k = нормализовать(ключ)
    if not k:
        return False
    if not re.fullmatch(r"[0-9a-zа-яё]+", k, flags=re.IGNORECASE):
        return False
    if k in нормализовать(короткое_имя):
        return True
    for часть in разобрать_части_имени(короткое_имя):
        for вариант in варианты_токена(часть):
            if нормализовать(вариант) == k:
                return True
    return False


def очистить_ключи_от_чужих(ключи: КлючиСтроки, короткое_имя: str) -> КлючиСтроки:
    def filtr(items: list[str]) -> list[str]:
        return [x for x in items if ключ_обоснован_именем(короткое_имя, x)]

    return КлючиСтроки(
        and_full=filtr(ключи.and_full),
        and_not=filtr(ключи.and_not),
        and_non=filtr(ключи.and_non),
        or_full=filtr(ключи.or_full),
        or_not=filtr(ключи.or_not),
        or_non=filtr(ключи.or_non),
        дубликат=ключи.дубликат,
        комментарий=ключи.комментарий,
    )


def извлечь_токены_и_буквы(
    короткое_имя: str,
    мин_длина: int,
    резервный: bool = False,
) -> tuple[list[str], list[str]]:
    части = разобрать_части_имени(короткое_имя)
    стоп = СТОП_СЛОВА if not резервный else {"гк", "группа", "холдинг", "ооо", "ао", "зао", "пао"}
    общие = ОБЩИЕ_ТОКЕНЫ if not резервный else set()

    токены: list[str] = []
    буквы: list[str] = []
    for часть in части:
        if len(часть) == 1 and часть.isalpha():
            буквы.append(часть)
            continue
        if len(часть) < мин_длина:
            continue
        if часть.lower() in стоп:
            continue
        if часть.lower() in общие:
            continue
        токены.append(часть)

    def uniq(items: list[str]) -> list[str]:
        out: list[str] = []
        seen: set[str] = set()
        for т in items:
            кл = т.lower()
            if кл not in seen:
                seen.add(кл)
                out.append(т)
        return out

    return uniq(токены), uniq(буквы)


def частота_токенов(записи: list[ЗаписьГСЗ]) -> dict[str, int]:
    частоты: dict[str, int] = {}
    for запись in записи:
        норм = нормализовать(запись.короткое_имя)
        for т in запись.токены:
            кл = т.lower()
            if кл in норм:
                частоты[кл] = частоты.get(кл, 0) + 1
    return частоты


def отсортировать_токены(токены: list[str], частоты: dict[str, int]) -> list[str]:
    return sorted(токены, key=lambda т: (частоты.get(т.lower(), 9999), -len(т), т.lower()))


def выбрать_режим_токена(токен: str, порог_not: int) -> bool:
    if len(токен) <= порог_not and токен.isalpha():
        return False
    return True


def лимиты_колонок(колонки: КолонкиКлючей) -> dict[str, int]:
    return {
        "and_full": len(колонки.and_full),
        "and_not": len(колонки.and_not),
        "and_non": len(колонки.and_non),
        "or_full": len(колонки.or_full),
        "or_not": len(колонки.or_not),
        "or_non": len(колонки.or_non),
    }


def построить_ключи_из_and(
    токены: list[tuple[str, bool]],
    лимиты: dict[str, int],
) -> КлючиСтроки | None:
    ключи = КлючиСтроки()
    for слово, is_full in токены:
        if is_full:
            if len(ключи.and_full) >= лимиты["and_full"]:
                return None
            ключи.and_full.append(слово)
        else:
            if len(ключи.and_not) >= лимиты["and_not"]:
                return None
            ключи.and_not.append(слово)
    return ключи


def построить_индекс_по_коротким_именам(записи: list[ЗаписьГСЗ]) -> list[str]:
    return [нормализовать(з.короткое_имя) for з in записи]


def кандидаты_коллизий(
    ключи: КлючиСтроки,
    своя: ЗаписьГСЗ,
    все: list[ЗаписьГСЗ],
    норм_имена: list[str],
) -> list[ЗаписьГСЗ]:
    and_слова = [нормализовать(w) for w in (ключи.and_full + ключи.and_not) if w]
    if not and_слова:
        return [з for з in все if з is not своя]
    якорь = max(and_слова, key=len)
    out: list[ЗаписьГСЗ] = []
    for i, з in enumerate(все):
        if з is своя:
            continue
        текст = норм_имена[i]
        if якорь not in текст:
            continue
        if any(w not in текст for w in and_слова):
            continue
        out.append(з)
    return out


def уникально_для_записи(
    ключи: КлючиСтроки,
    своя: ЗаписьГСЗ,
    все: list[ЗаписьГСЗ],
    норм_имена: list[str] | None = None,
) -> bool:
    if not строка_совпала(своя.короткое_имя, ключи):
        return False
    if норм_имена is None:
        норм_имена = построить_индекс_по_коротким_именам(все)
    for другая in кандидаты_коллизий(ключи, своя, все, норм_имена):
        if строка_совпала(другая.короткое_имя, ключи):
            return False
    return True


def добавить_or_not_из_букв(
    ключи: КлючиСтроки,
    буквы: list[str],
    лимиты: dict[str, int],
    короткое_имя: str,
) -> КлючиСтроки:
    """«Строй-А» → or_not: А, A."""
    if лимиты["or_not"] <= 0 or not буквы:
        return ключи
    out = КлючиСтроки(
        and_full=list(ключи.and_full),
        and_not=list(ключи.and_not),
        and_non=[],
        or_full=list(ключи.or_full),
        or_not=list(ключи.or_not),
        or_non=[],
        дубликат=ключи.дубликат,
        комментарий=ключи.комментарий,
    )
    seen = {нормализовать(x) for x in out.or_not}
    for буква in буквы:
        for вариант in варианты_токена(буква):
            if len(out.or_not) >= лимиты["or_not"]:
                return очистить_ключи_от_чужих(out, короткое_имя)
            кл = нормализовать(вариант)
            if кл in seen:
                continue
            if not ключ_обоснован_именем(короткое_имя, вариант):
                continue
            out.or_not.append(вариант)
            seen.add(кл)
    return очистить_ключи_от_чужих(out, короткое_имя)


def попробовать_or_дополнение(
    ключи: КлючиСтроки,
    своя: ЗаписьГСЗ,
    кандидаты: list[str],
    порог_not: int,
    лимиты: dict[str, int],
) -> КлючиСтроки:
    if ключи.and_full or ключи.and_not:
        return ключи
    if лимиты["or_full"] <= 0 and лимиты["or_not"] <= 0:
        return ключи
    for т in кандидаты:
        if not ключ_обоснован_именем(своя.короткое_имя, т):
            continue
        is_full = выбрать_режим_токена(т, порог_not)
        пробные = КлючиСтроки(or_full=list(ключи.or_full), or_not=list(ключи.or_not))
        if is_full:
            if len(пробные.or_full) >= лимиты["or_full"]:
                continue
            пробные.or_full.append(т)
        else:
            if len(пробные.or_not) >= лимиты["or_not"]:
                continue
            пробные.or_not.append(т)
        пробные = очистить_ключи_от_чужих(пробные, своя.короткое_имя)
        if строка_совпала(своя.короткое_имя, пробные):
            return пробные
    return ключи


def подобрать_ключи(
    запись: ЗаписьГСЗ,
    все: list[ЗаписьГСЗ],
    конфиг: dict[str, Any],
    частоты: dict[str, int],
    колонки: КолонкиКлючей,
    норм_имена: list[str],
) -> КлючиСтроки:
    порог_not = int(конфиг["short_token_not_max_len"])
    лимиты = лимиты_колонок(колонки)
    max_and = лимиты["and_full"] + лимиты["and_not"]
    имя = запись.короткое_имя

    кандидаты = [т for т in отсортировать_токены(запись.токены, частоты) if ключ_обоснован_именем(имя, т)]
    if not кандидаты:
        резерв, буквы_рез = извлечь_токены_и_буквы(имя, int(конфиг["min_token_length"]), резервный=True)
        if not запись.буквы:
            запись.буквы = буквы_рез
        кандидаты = [т for т in отсортировать_токены(резерв, частоты) if ключ_обоснован_именем(имя, т)]

    if not кандидаты and not запись.буквы:
        return КлючиСтроки(дубликат=True, комментарий="нет токенов для ключа")

    варианты = [(т, выбрать_режим_токена(т, порог_not)) for т in кандидаты]

    def с_буквами(к: КлючиСтроки) -> КлючиСтроки:
        return добавить_or_not_из_букв(к, запись.буквы, лимиты, имя)

    for размер in range(1, (min(max_and, len(варианты)) + 1) if варианты else 1):
        if not варианты:
            break
        for комбо in combinations(варианты, размер):
            ключи = построить_ключи_из_and(list(комбо), лимиты)
            if ключи is None:
                continue
            ключи = с_буквами(очистить_ключи_от_чужих(ключи, имя))
            if уникально_для_записи(ключи, запись, все, норм_имена):
                return ключи

    top = кандидаты[: max(max_and + 2, 4)]
    for размер in range(1, (min(max_and, len(top)) + 1) if top else 1):
        if not top:
            break
        for индексы in combinations(range(len(top)), размер):
            for маска in range(2 ** len(индексы)):
                комбо = [(top[idx], bool(маска & (1 << бит))) for бит, idx in enumerate(индексы)]
                ключи = построить_ключи_из_and(комбо, лимиты)
                if ключи is None:
                    continue
                ключи = с_буквами(очистить_ключи_от_чужих(ключи, имя))
                if уникально_для_записи(ключи, запись, все, норм_имена):
                    return ключи

    комбо_base = [(т, выбрать_режим_токена(т, порог_not)) for т in кандидаты[:max_and]]
    ключи = построить_ключи_из_and(комбо_base, лимиты) or КлючиСтроки()
    ключи = очистить_ключи_от_чужих(ключи, имя)
    if not строка_совпала(имя, ключи):
        for т in кандидаты:
            проб = построить_ключи_из_and([(т, True)], лимиты)
            if проб:
                проб = очистить_ключи_от_чужих(проб, имя)
                if строка_совпала(имя, проб):
                    ключи = проб
                    break

    ключи = попробовать_or_дополнение(ключи, запись, кандидаты, порог_not, лимиты)
    ключи = с_буквами(ключи)
    ключи = очистить_ключи_от_чужих(ключи, имя)

    if уникально_для_записи(ключи, запись, все, норм_имена):
        return ключи

    ключи.дубликат = True
    ключи.комментарий = "не удалось подобрать уникальные ключи"
    return ключи


def прочитать_имена_из_excel(путь: Path, table_name: str, gsz_column: str) -> list[str]:
    rows = read_excel_table(путь, table_name)
    if not rows:
        return []
    if gsz_column not in rows[0]:
        raise ValueError(
            f"В таблице '{table_name}' нет колонки '{gsz_column}'. Доступны: {list(rows[0].keys())}"
        )
    имена: list[str] = []
    for row in rows:
        raw = row.get(gsz_column)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            имена.append(text)
    return имена


def отметить_групповые_дубликаты(записи: list[ЗаписьГСЗ], колонки: КолонкиКлючей) -> None:
    пустой = {к: "" for к in колонки.все()}
    for i, а in enumerate(записи):
        for б in записи[i + 1 :]:
            da = а.ключи.в_словарь(колонки)
            db = б.ключи.в_словарь(колонки)
            if da == db and da != пустой:
                а.ключи.дубликат = True
                б.ключи.дубликат = True
                if not а.ключи.комментарий:
                    а.ключи.комментарий = "одинаковые ключи с другой строкой"
                if not б.ключи.комментарий:
                    б.ключи.комментарий = "одинаковые ключи с другой строкой"
    for i, а in enumerate(записи):
        for б in записи[i + 1 :]:
            if строка_совпала(а.короткое_имя, б.ключи) and строка_совпала(а.короткое_имя, а.ключи):
                а.ключи.дубликат = True
                б.ключи.дубликат = True
            if строка_совпала(б.короткое_имя, а.ключи) and строка_совпала(б.короткое_имя, б.ключи):
                а.ключи.дубликат = True
                б.ключи.дубликат = True


def сохранить_excel(
    записи: list[ЗаписьГСЗ],
    путь: Path,
    gsz_column: str,
    колонки: КолонкиКлючей,
) -> None:
    путь.parent.mkdir(parents=True, exist_ok=True)
    key_cols = колонки.все()
    заголовки = [gsz_column] + key_cols + ["ключи_задублированы", "комментарий_ключей"]
    книга = Workbook()
    лист = книга.active
    лист.title = "_base_gsz"
    лист.append(заголовки)
    for запись in записи:
        словарь = запись.ключи.в_словарь(колонки)
        строка = [запись.полное_имя]
        строка.extend(словарь.get(к, "") for к in key_cols)
        строка.append("да" if запись.ключи.дубликат else "")
        строка.append(запись.ключи.комментарий)
        лист.append(строка)
    ref = f"A1:{get_column_letter(len(заголовки))}{len(записи) + 1}"
    таблица = Table(displayName="_base_gsz", ref=ref)
    таблица.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    лист.add_table(таблица)
    книга.save(путь)


def проверить_все_ключи_из_имени(записи: list[ЗаписьГСЗ]) -> int:
    """Сколько ключей нарушают правило «только из своего имени»."""
    нарушений = 0
    for з in записи:
        for поле in (
            з.ключи.and_full,
            з.ключи.and_not,
            з.ключи.and_non,
            з.ключи.or_full,
            з.ключи.or_not,
            з.ключи.or_non,
        ):
            for k in поле:
                if not ключ_обоснован_именем(з.короткое_имя, k):
                    нарушений += 1
    return нарушений


def main() -> None:
    настроить_логирование()
    конфиг = загрузить_конфиг()
    config_dir = Path(str(конфиг["_config_dir"])).resolve()
    колонки: КолонкиКлючей = конфиг["колонки"]
    вход = resolve_path(str(конфиг["input_xlsx"]), config_dir)
    выход = resolve_path(str(конфиг["output_file"]), config_dir)
    if конфиг.get("output_add_timestamp", True):
        выход = with_timestamp_suffix(выход, str(конфиг["output_timestamp_format"]))
    base_table = str(конфиг["base_table"])
    gsz_column = str(конфиг["gsz_column"])
    мин_длина = int(конфиг["min_token_length"])

    logging.info("Источник: %s | таблица=%s | колонка=%s", вход, base_table, gsz_column)
    logging.info("Ключевые колонки: %s", ", ".join(колонки.все()))

    полные_имена = прочитать_имена_из_excel(вход, base_table, gsz_column)
    logging.info("Загружено строк: %d", len(полные_имена))

    записи: list[ЗаписьГСЗ] = []
    for полное in полные_имена:
        короткое = полное.split(",", 1)[0].strip()
        токены, буквы = извлечь_токены_и_буквы(короткое, мин_длина)
        записи.append(ЗаписьГСЗ(полное_имя=полное, короткое_имя=короткое, токены=токены, буквы=буквы))

    # Контрольный пример «Строй-А»
    демо_т, демо_б = извлечь_токены_и_буквы("Строй-А", мин_длина)
    logging.info("Проверка разбора «Строй-А»: токены=%s буквы=%s варианты=%s", демо_т, демо_б, варианты_токена("А"))

    logging.info("Подбор ключей (только из своего имени + транслит)...")
    частоты = частота_токенов(записи)
    норм_имена = построить_индекс_по_коротким_именам(записи)
    for idx, запись in enumerate(записи, start=1):
        запись.ключи = подобрать_ключи(запись, записи, конфиг, частоты, колонки, норм_имена)
        if idx % 500 == 0 or idx == len(записи):
            logging.info("Прогресс подбора: %d/%d", idx, len(записи))

    отметить_групповые_дубликаты(записи, колонки)

    нарушений = проверить_все_ключи_из_имени(записи)
    дубликатов = sum(1 for з in записи if з.ключи.дубликат)
    без_токенов = sum(1 for з in записи if not з.токены and not з.буквы)
    с_or_not = sum(1 for з in записи if з.ключи.or_not)
    с_non = sum(1 for з in записи if з.ключи.and_non or з.ключи.or_non)
    уникальных = len(записи) - дубликатов

    сохранить_excel(записи, выход, gsz_column, колонки)

    logging.info("Сохранено: %s", выход)
    logging.info("Уникальных ключей: %d", уникальных)
    logging.info("Строк с пометкой дубликат: %d", дубликатов)
    logging.info("Строк с or_not (в т.ч. буквы): %d", с_or_not)
    logging.info("Строк с non-ключами: %d (ожидаемо 0 — non вручную)", с_non)
    logging.info("Строк без токенов: %d", без_токенов)
    logging.info("Нарушений «ключ не из имени»: %d", нарушений)


if __name__ == "__main__":
    main()
