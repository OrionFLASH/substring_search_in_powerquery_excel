#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ERZ JSON → Excel: лист групп и лист уникальных компаний по ИНН.

Читает результат DevTools-скрапера (ERZ_Full_*.json) из input/,
формирует два листа с форматированием по блоку erz_json_to_excel в config.json.
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

КОРЕНЬ: Path = Path(__file__).resolve().parent.parent
КОНФИГ_ПО_УМОЛЧАНИЮ: Path = КОРЕНЬ / "config.json"

logger: logging.Logger = logging.getLogger("erz_json_to_excel")

GROUP_COLUMN_KEYS: tuple[str, ...] = (
    "group_id",
    "group_name",
    "group_url_id",
    "regions",
    "region_count",
    "group_companies",
    "group_companies_count",
    "brand_companies",
    "brand_companies_count",
)

COMPANY_COLUMN_KEYS: tuple[str, ...] = (
    "inn",
    "company_name",
    "ogrn",
    "url_id",
    "role",
    "groups",
    "groups_count",
    "group_regions",
    "group_regions_count",
)


def настроить_логирование(
    logs_dir: Path,
    log_file_prefix: str,
    log_to_file: bool,
) -> Path | None:
    """Консоль + опциональный файл лога."""
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter(
        "%(asctime)s - [%(levelname)s] - %(message)s [def: %(funcName)s]",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(fmt)
    logger.addHandler(console)

    if not log_to_file:
        return None

    logs_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H")
    log_path = logs_dir / f"{log_file_prefix}_{stamp}.log"
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)
    return log_path


def загрузить_конфиг(path: Path) -> dict[str, Any]:
    """Чтение JSON-конфига."""
    if not path.exists():
        raise FileNotFoundError(f"config-json не найден: {path}")
    with path.open(encoding="utf-8") as f:
        data: dict[str, Any] = json.load(f)
    return data


def разрешить_путь(value: str, base_dir: Path) -> Path:
    """Относительный путь — от каталога config.json."""
    p = Path(value).expanduser()
    if p.is_absolute():
        return p.resolve()
    return (base_dir / p).resolve()


def с_таймштампом(path: Path, pattern: str = "%Y%m%d_%H%M%S") -> Path:
    """Добавить таймштамп перед расширением файла."""
    ts = datetime.now().strftime(pattern)
    if path.suffix:
        return path.with_name(f"{path.stem}_{ts}{path.suffix}")
    return path.with_name(f"{path.name}_{ts}")


def загрузить_erz_json(path: Path) -> dict[str, Any]:
    """Прочитать ERZ_Full JSON."""
    with path.open(encoding="utf-8") as f:
        data: dict[str, Any] = json.load(f)
    if not isinstance(data, dict):
        raise ValueError(f"Ожидался объект JSON в корне: {path}")
    if "groups" not in data or not isinstance(data["groups"], list):
        raise ValueError(f"В JSON нет массива groups: {path}")
    logger.info("Загружен JSON: %s (групп: %s)", path, len(data["groups"]))
    return data


def найти_входной_json(
    input_json: str | None,
    input_dir: Path,
    input_glob: str,
    config_dir: Path,
) -> Path:
    """Явный файл либо самый новый по glob в input_dir."""
    if input_json:
        path = разрешить_путь(input_json, config_dir)
        if not path.exists():
            raise FileNotFoundError(f"Входной JSON не найден: {path}")
        return path

    if not input_dir.is_dir():
        raise FileNotFoundError(f"Каталог input не найден: {input_dir}")

    candidates = sorted(
        input_dir.glob(input_glob),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"Нет файлов по шаблону «{input_glob}» в {input_dir}"
        )
    logger.info("Выбран JSON по glob: %s", candidates[0].name)
    return candidates[0]


def формат_компании(company: dict[str, Any], template: str) -> str:
    """Одна строка компании по шаблону {name} {inn} {ogrn} {urlId} {id}."""
    return template.format(
        name=str(company.get("name") or ""),
        inn=str(company.get("inn") or ""),
        ogrn=str(company.get("ogrn") or ""),
        urlId=str(company.get("urlId") or ""),
        id=str(company.get("id") or ""),
    )


def склеить_список(items: list[str], joiner: str) -> str:
    """Уникальные непустые значения с сохранением порядка."""
    seen: OrderedDict[str, None] = OrderedDict()
    for item in items:
        text = str(item).strip()
        if text and text not in seen:
            seen[text] = None
    return joiner.join(seen.keys())


def карта_регионов_группы(group: dict[str, Any]) -> dict[str, str]:
    """regionKey → название региона."""
    out: dict[str, str] = {}
    for reg in group.get("regions") or []:
        if not isinstance(reg, dict):
            continue
        key = str(reg.get("regionKey") or "").strip()
        name = str(reg.get("region") or "").strip()
        if key and name:
            out[key] = name
    return out


def регионы_компании_в_группе(
    company: dict[str, Any],
    group: dict[str, Any],
) -> list[str]:
    """Названия регионов присутствия компании внутри группы."""
    key_to_name = карта_регионов_группы(group)
    names: list[str] = []
    for loc in company.get("locations") or []:
        if not isinstance(loc, dict):
            continue
        rk = str(loc.get("regionKey") or "").strip()
        if rk and rk in key_to_name:
            names.append(key_to_name[rk])
    if names:
        return names
    # Нет regionKey у локаций (часто brandCompanies) — регионы группы.
    return [str(r.get("region") or "").strip() for r in (group.get("regions") or []) if isinstance(r, dict)]


def списки_компаний_группы(
    group: dict[str, Any],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], bool]:
    """groupCompanies, brandCompanies и флаг names-only (нет JOIN, есть names)."""
    group_cos = [c for c in (group.get("groupCompanies") or []) if isinstance(c, dict)]
    brand_cos = [c for c in (group.get("brandCompanies") or []) if isinstance(c, dict)]
    names_only = not group_cos and bool(brand_cos)
    return group_cos, brand_cos, names_only


def построить_лист_групп(
    data: dict[str, Any],
    joiner: str,
    company_template: str,
) -> list[dict[str, Any]]:
    """Строки листа «группы застройщиков»."""
    rows: list[dict[str, Any]] = []
    names_only_n = 0
    for group in data.get("groups") or []:
        if not isinstance(group, dict):
            continue
        region_names = [
            str(r.get("region") or "").strip()
            for r in (group.get("regions") or [])
            if isinstance(r, dict) and str(r.get("region") or "").strip()
        ]
        group_cos, brand_cos, names_only = списки_компаний_группы(group)
        # JSON без JOIN: names = полный список → дублируем в колонку «компании группы»
        group_for_sheet = brand_cos if names_only else group_cos
        if names_only:
            names_only_n += 1
        rows.append(
            {
                "group_id": str(group.get("id") or ""),
                "group_name": str(group.get("name") or ""),
                "group_url_id": str(group.get("urlId") or ""),
                "regions": склеить_список(region_names, joiner),
                "region_count": len(OrderedDict.fromkeys(region_names)),
                "group_companies": склеить_список(
                    [формат_компании(c, company_template) for c in group_for_sheet],
                    joiner,
                ),
                "group_companies_count": len(group_for_sheet),
                "brand_companies": склеить_список(
                    [формат_компании(c, company_template) for c in brand_cos],
                    joiner,
                ),
                "brand_companies_count": len(brand_cos),
            }
        )
    if names_only_n:
        logger.info(
            "Лист групп: строк %s (без JOIN, список из names: %s групп)",
            len(rows),
            names_only_n,
        )
    else:
        logger.info("Лист групп: строк %s", len(rows))
    return rows


def построить_лист_компаний(
    data: dict[str, Any],
    joiner: str,
    group_region_template: str,
    role_labels: dict[str, str],
) -> list[dict[str, Any]]:
    """Уникальные компании по ИНН: роль, группы, группы+регионы."""
    # inn → агрегат
    by_inn: OrderedDict[str, dict[str, Any]] = OrderedDict()

    def ensure(inn: str, company: dict[str, Any]) -> dict[str, Any]:
        if inn not in by_inn:
            by_inn[inn] = {
                "inn": inn,
                "names": OrderedDict(),
                "ogrns": OrderedDict(),
                "url_ids": OrderedDict(),
                "in_group": False,
                "in_brand": False,
                "group_names": OrderedDict(),
                "group_regions": OrderedDict(),
            }
        item = by_inn[inn]
        name = str(company.get("name") or "").strip()
        if name:
            item["names"][name] = None
        ogrn = str(company.get("ogrn") or "").strip()
        if ogrn:
            item["ogrns"][ogrn] = None
        url_id = str(company.get("urlId") or "").strip()
        if url_id:
            item["url_ids"][url_id] = None
        return item

    for group in data.get("groups") or []:
        if not isinstance(group, dict):
            continue
        group_name = str(group.get("name") or "").strip() or str(group.get("id") or "")
        group_cos, brand_cos, names_only = списки_компаний_группы(group)

        for company in group_cos:
            inn = str(company.get("inn") or "").strip()
            if not inn:
                continue
            item = ensure(inn, company)
            item["in_group"] = True
            item["group_names"][group_name] = None
            for region_name in регионы_компании_в_группе(company, group):
                if not region_name:
                    continue
                label = group_region_template.format(
                    group=group_name,
                    region=region_name,
                )
                item["group_regions"][label] = None

        for company in brand_cos:
            inn = str(company.get("inn") or "").strip()
            if not inn:
                continue
            item = ensure(inn, company)
            # names-only: names ≈ полный список компаний группы
            if names_only:
                item["in_group"] = True
            else:
                item["in_brand"] = True
            item["group_names"][group_name] = None
            for region_name in регионы_компании_в_группе(company, group):
                if not region_name:
                    continue
                label = group_region_template.format(
                    group=group_name,
                    region=region_name,
                )
                item["group_regions"][label] = None

    role_group = str(role_labels.get("group", "компания группы"))
    role_brand = str(role_labels.get("brand", "компания бренда"))
    role_both = str(role_labels.get("both", "группа и бренд"))

    rows: list[dict[str, Any]] = []
    for inn, item in by_inn.items():
        if item["in_group"] and item["in_brand"]:
            role = role_both
        elif item["in_group"]:
            role = role_group
        else:
            role = role_brand

        group_list = list(item["group_names"].keys())
        region_list = list(item["group_regions"].keys())
        rows.append(
            {
                "inn": inn,
                "company_name": joiner.join(item["names"].keys()),
                "ogrn": joiner.join(item["ogrns"].keys()),
                "url_id": joiner.join(item["url_ids"].keys()),
                "role": role,
                "groups": joiner.join(group_list),
                "groups_count": len(group_list),
                "group_regions": joiner.join(region_list),
                "group_regions_count": len(region_list),
            }
        )

    logger.info("Лист компаний: уникальных ИНН %s", len(rows))
    return rows


def применить_форматирование(
    ws: Any,
    format_cfg: dict[str, Any],
    column_keys: list[str],
    column_specs: dict[str, dict[str, Any]],
) -> None:
    """Заголовки, автофильтр, freeze, ширины и выравнивание."""
    header_bold = bool(format_cfg.get("header_bold", True))
    header_center = bool(format_cfg.get("header_center", True))
    header_wrap = bool(format_cfg.get("header_wrap", True))
    freeze_rows = max(0, int(format_cfg.get("freeze_rows", 1)))
    freeze_cols = max(0, int(format_cfg.get("freeze_cols", 1)))
    data_v_center = bool(format_cfg.get("data_vertical_center", True))
    data_h_default = str(format_cfg.get("data_horizontal", "left"))

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.alignment = Alignment(
                horizontal="center" if header_center else "general",
                vertical="center",
                wrap_text=header_wrap,
            )
            if header_bold:
                cell.font = Font(bold=True)

    if ws.max_row >= 2:
        for col_idx, key in enumerate(column_keys, start=1):
            spec = column_specs.get(key, {})
            wrap = bool(spec.get("wrap", True))
            align_h = str(spec.get("align", data_h_default))
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(
                    horizontal=align_h,
                    vertical="center" if data_v_center else "bottom",
                    wrap_text=wrap,
                )

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    if freeze_rows > 0 or freeze_cols > 0:
        ws.freeze_panes = ws.cell(row=freeze_rows + 1, column=freeze_cols + 1)

    for col_idx, key in enumerate(column_keys, start=1):
        width = float(column_specs.get(key, {}).get("width", 30.0))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def записать_лист(
    wb: Workbook,
    sheet_name: str,
    rows: list[dict[str, Any]],
    column_keys: list[str],
    output_columns: dict[str, str],
    format_cfg: dict[str, Any],
    *,
    is_first: bool,
) -> None:
    """Добавить лист с данными и форматированием."""
    title = (sheet_name or "Sheet")[:31]
    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    headers = [str(output_columns.get(k, k)) for k in column_keys]
    column_specs: dict[str, dict[str, Any]] = dict(format_cfg.get("columns") or {})
    ws.append(headers)
    for row in rows:
        ws.append([row.get(k, "") for k in column_keys])
    применить_форматирование(ws, format_cfg, column_keys, column_specs)


def сохранить_excel(
    output_path: Path,
    group_rows: list[dict[str, Any]],
    company_rows: list[dict[str, Any]],
    settings: dict[str, Any],
) -> None:
    """Записать книгу с двумя листами."""
    wb = Workbook()
    записать_лист(
        wb,
        str(settings["groups_sheet"]),
        group_rows,
        list(GROUP_COLUMN_KEYS),
        dict(settings["groups_columns"]),
        dict(settings["groups_format"]),
        is_first=True,
    )
    записать_лист(
        wb,
        str(settings["companies_sheet"]),
        company_rows,
        list(COMPANY_COLUMN_KEYS),
        dict(settings["companies_columns"]),
        dict(settings["companies_format"]),
        is_first=False,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "Сохранено: %s (групп=%s, компаний=%s)",
        output_path,
        len(group_rows),
        len(company_rows),
    )


def разрешить_настройки(args: argparse.Namespace) -> dict[str, Any]:
    """Собрать настройки из config + CLI."""
    config_path = Path(args.config_json).expanduser().resolve()
    cfg = загрузить_конфиг(config_path)
    block: dict[str, Any] = dict(cfg.get("erz_json_to_excel") or {})

    def pick(cli_val: Any, key: str, default: Any = None) -> Any:
        if cli_val is not None:
            return cli_val
        if key in block and block[key] is not None:
            return block[key]
        return default

    settings: dict[str, Any] = {
        "_config_dir": str(config_path.parent),
        "input_json": pick(args.input_json, "input_json"),
        "input_dir": str(pick(None, "input_dir", "input")),
        "input_glob": str(pick(None, "input_glob", "ERZ_Full_*.json")),
        "output_xlsx": pick(args.output_xlsx, "output_xlsx", "output/erz_from_json.xlsx"),
        "output_add_timestamp": bool(
            pick(args.output_add_timestamp, "output_add_timestamp", True)
        ),
        "output_timestamp_format": str(
            pick(None, "output_timestamp_format", "%Y%m%d_%H%M%S")
        ),
        "groups_sheet": str(pick(None, "groups_sheet", "_ERZ_GROUPS")),
        "companies_sheet": str(pick(None, "companies_sheet", "_ERZ_COMPANIES")),
        "multi_value_joiner": str(pick(None, "multi_value_joiner", ";\n")),
        "company_line_template": str(
            pick(None, "company_line_template", "{name} (ИНН {inn})")
        ),
        "group_region_template": str(
            pick(None, "group_region_template", "{group}, {region}")
        ),
        "role_labels": dict(
            pick(
                None,
                "role_labels",
                {
                    "group": "компания группы",
                    "brand": "компания бренда",
                    "both": "группа и бренд",
                },
            )
        ),
        "groups_columns": dict(pick(None, "groups_columns", {})),
        "companies_columns": dict(pick(None, "companies_columns", {})),
        "groups_format": dict(pick(None, "groups_format", {})),
        "companies_format": dict(pick(None, "companies_format", {})),
        "log_to_file": bool(pick(None, "log_to_file", True)),
        "logs_dir": str(pick(None, "logs_dir", "log")),
        "log_file_prefix": str(pick(None, "log_file_prefix", "INFO_erz_json_to_excel")),
    }
    return settings


def make_arg_parser() -> argparse.ArgumentParser:
    """CLI."""
    parser = argparse.ArgumentParser(
        description="ERZ JSON → Excel (группы и уникальные компании)"
    )
    parser.add_argument(
        "--config-json",
        default=str(КОНФИГ_ПО_УМОЛЧАНИЮ),
        help="Путь к config.json",
    )
    parser.add_argument(
        "--input-json",
        default=None,
        help="Явный путь к ERZ_Full JSON (иначе glob в input_dir)",
    )
    parser.add_argument("--output-xlsx", default=None, help="Выходной Excel")
    parser.add_argument(
        "--output-add-timestamp",
        default=None,
        type=lambda v: str(v).lower() in {"1", "true", "yes", "да"},
        help="Добавить таймштамп к имени выходного файла",
    )
    return parser


def main() -> None:
    """Точка входа."""
    parser = make_arg_parser()
    args = parser.parse_args()
    settings = разрешить_настройки(args)
    config_dir = Path(str(settings["_config_dir"])).resolve()

    log_path = настроить_логирование(
        logs_dir=разрешить_путь(str(settings["logs_dir"]), config_dir),
        log_file_prefix=str(settings["log_file_prefix"]),
        log_to_file=bool(settings["log_to_file"]),
    )
    if log_path is not None:
        logger.info("Файл лога: %s", log_path)

    input_dir = разрешить_путь(str(settings["input_dir"]), config_dir)
    input_path = найти_входной_json(
        input_json=settings["input_json"],
        input_dir=input_dir,
        input_glob=str(settings["input_glob"]),
        config_dir=config_dir,
    )

    output_xlsx = разрешить_путь(str(settings["output_xlsx"]), config_dir)
    if settings["output_add_timestamp"]:
        output_xlsx = с_таймштампом(
            output_xlsx, str(settings["output_timestamp_format"])
        )

    data = загрузить_erz_json(input_path)
    joiner = str(settings["multi_value_joiner"])

    group_rows = построить_лист_групп(
        data,
        joiner=joiner,
        company_template=str(settings["company_line_template"]),
    )
    company_rows = построить_лист_компаний(
        data,
        joiner=joiner,
        group_region_template=str(settings["group_region_template"]),
        role_labels=dict(settings["role_labels"]),
    )

    сохранить_excel(output_xlsx, group_rows, company_rows, settings)
    logger.info("Готово.")


if __name__ == "__main__":
    main()
